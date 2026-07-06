from flask import Flask, render_template, request, redirect, url_for, session, flash, send_from_directory, send_file, abort, jsonify
import sqlite3, os, uuid, secrets, json, tempfile, time, re
from datetime import datetime, timedelta
from io import BytesIO
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from markupsafe import Markup, escape
import openpyxl
import qrcode
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_NAME = os.environ.get('SQLITE_DB_PATH', os.path.join(BASE_DIR, 'tournament_events.db'))
DATABASE_URL = os.environ.get('DATABASE_URL', '').strip()
if DATABASE_URL.startswith('postgres://'):
    DATABASE_URL = 'postgresql://' + DATABASE_URL[len('postgres://'):]
if DATABASE_URL.startswith('postgresql+psycopg://'):
    DATABASE_URL = 'postgresql://' + DATABASE_URL[len('postgresql+psycopg://'):]
IS_POSTGRES = DATABASE_URL.startswith('postgresql://')
UPLOAD_FOLDER = os.environ.get('UPLOAD_FOLDER', os.path.join(BASE_DIR, 'static', 'uploads'))
ALLOWED_EXTENSIONS = {'png','jpg','jpeg','pdf','webp'}
CERT_IMAGE_EXTENSIONS = {'png','jpg','jpeg','webp'}
EXCEL_EXTENSIONS = {'xlsx'}
app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'change-this-secret-key')
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 15 * 1024 * 1024
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
BULK_IMPORT_CACHE_FOLDER = os.environ.get('BULK_IMPORT_CACHE_FOLDER', os.path.join(tempfile.gettempdir(), 'register_team_bulk_imports'))
os.makedirs(BULK_IMPORT_CACHE_FOLDER, exist_ok=True)


def format_signer_position(value):
    """แสดงตำแหน่งผู้ลงนามให้ขึ้นบรรทัดใหม่ได้ทั้งจาก Enter และการเว้นวรรค 2 ครั้ง"""
    text = str(value or '').replace('\r\n', '\n').replace('\r', '\n').strip()
    text = re.sub(r' {2,}', '\n', text)
    lines = [escape(line.strip()) for line in text.split('\n')]
    return Markup('<br>'.join(lines))


THAI_MONTHS_FULL = [
    '', 'มกราคม', 'กุมภาพันธ์', 'มีนาคม', 'เมษายน', 'พฤษภาคม', 'มิถุนายน',
    'กรกฎาคม', 'สิงหาคม', 'กันยายน', 'ตุลาคม', 'พฤศจิกายน', 'ธันวาคม'
]


def thai_today_text():
    """วันที่ปัจจุบันตามเวลาไทย สำหรับใช้เมื่อแอดมินยังไม่ได้กรอกช่วงวันจัดกิจกรรม"""
    d = datetime.utcnow() + timedelta(hours=7)
    return f"{d.day} {THAI_MONTHS_FULL[d.month]} {d.year + 543}"


def format_cert_issue_date(value):
    """
    แปลงช่องวันจัดกิจกรรมเป็นวันที่ในบรรทัด "ให้ไว้ ณ วันที่ ..."
    - ถ้าเป็นช่วง เช่น 8-10 กรกฎาคม 2569 ให้ใช้วันสุดท้าย คือ 10 กรกฎาคม 2569
    - ถ้าระบุวันเดียว ให้ใช้วันนั้น
    - ถ้าไม่ได้กรอก ให้ใช้วันที่ปัจจุบันตามเวลาไทย
    """
    text = str(value or '').replace('\r', ' ').replace('\n', ' ').strip()
    if not text:
        return thai_today_text()

    text = re.sub(r'\s+', ' ', text)
    text = re.sub(r'^วันที่\s*', '', text).strip()
    text = text.replace('–', '-').replace('—', '-').replace('−', '-')

    # รูปแบบที่พบบ่อย: 8-10 กรกฎาคม 2569, 8 - 10 ก.ค. 2569, 8-10/7/2569
    m = re.search(r'(?:^|\D)(\d{1,2})\s*-\s*(\d{1,2})(.*)$', text)
    if m:
        return (m.group(2) + (m.group(3) or '')).strip()

    # รูปแบบอีกแบบ: 8 กรกฎาคม - 10 กรกฎาคม 2569
    m = re.search(r'-\s*(\d{1,2}\s+.+)$', text)
    if m:
        return m.group(1).strip()

    return text


app.jinja_env.filters['signer_position'] = format_signer_position
app.jinja_env.filters['cert_issue_date'] = format_cert_issue_date


class PostgresCursor:
    """ทำให้คำสั่ง SQL เดิมที่ใช้ ? ทำงานกับ psycopg ได้โดยไม่ต้องเขียน route ใหม่ทั้งหมด"""
    def __init__(self, cursor):
        self._cursor = cursor

    def execute(self, sql, params=()):
        self._cursor.execute(sql.replace('?', '%s'), params or ())
        return self

    def executemany(self, sql, params_seq):
        self._cursor.executemany(sql.replace('?', '%s'), params_seq)
        return self

    def fetchone(self):
        return self._cursor.fetchone()

    def fetchall(self):
        return self._cursor.fetchall()

    @property
    def rowcount(self):
        return self._cursor.rowcount

    def close(self):
        self._cursor.close()


class PostgresConnection:
    def __init__(self, connection):
        self._connection = connection

    def cursor(self):
        return PostgresCursor(self._connection.cursor())

    def execute(self, sql, params=()):
        return self.cursor().execute(sql, params)

    def commit(self):
        self._connection.commit()

    def rollback(self):
        self._connection.rollback()

    def close(self):
        self._connection.close()


def get_db():
    if IS_POSTGRES:
        import psycopg
        from psycopg.rows import dict_row
        return PostgresConnection(psycopg.connect(DATABASE_URL, row_factory=dict_row))
    conn = sqlite3.connect(DB_NAME, timeout=30)
    conn.row_factory = sqlite3.Row
    conn.execute('PRAGMA busy_timeout=30000')
    return conn


def ensure_column(c, table, name, ddl):
    if IS_POSTGRES:
        rows = c.execute('''SELECT column_name FROM information_schema.columns WHERE table_schema='public' AND table_name=?''', (table,)).fetchall()
        cols = [row['column_name'] for row in rows]
    else:
        cols = [row[1] for row in c.execute(f'PRAGMA table_info({table})').fetchall()]
    if name not in cols:
        c.execute(f'ALTER TABLE {table} ADD COLUMN {name} {ddl}')


def insert_returning_id(conn, sql, params):
    if IS_POSTGRES:
        return conn.execute(sql.rstrip().rstrip(';') + ' RETURNING id', params).fetchone()['id']
    c = conn.cursor(); c.execute(sql, params); return c.lastrowid


def init_db():
    conn = get_db(); c = conn.cursor()
    id_col = 'SERIAL PRIMARY KEY' if IS_POSTGRES else 'INTEGER PRIMARY KEY AUTOINCREMENT'
    c.execute(f'''CREATE TABLE IF NOT EXISTS users (id {id_col}, username TEXT UNIQUE NOT NULL, password TEXT NOT NULL, role TEXT NOT NULL DEFAULT 'admin')''')
    c.execute(f'''CREATE TABLE IF NOT EXISTS tournaments (id {id_col}, title TEXT NOT NULL, description TEXT, created_by INTEGER, created_at TEXT, is_open INTEGER NOT NULL DEFAULT 1, FOREIGN KEY(created_by) REFERENCES users(id))''')
    c.execute(f'''CREATE TABLE IF NOT EXISTS events (id {id_col}, tournament_id INTEGER NOT NULL, event_name TEXT, category_type TEXT NOT NULL, gender_type TEXT NOT NULL, age_group TEXT NOT NULL, max_slots INTEGER NOT NULL DEFAULT 0, fee INTEGER NOT NULL DEFAULT 0, team_size INTEGER NOT NULL DEFAULT 1, is_open INTEGER NOT NULL DEFAULT 1, created_at TEXT, FOREIGN KEY(tournament_id) REFERENCES tournaments(id))''')
    c.execute(f'''CREATE TABLE IF NOT EXISTS registrations (id {id_col}, event_id INTEGER NOT NULL, team_name TEXT, contact_name TEXT NOT NULL, phone TEXT NOT NULL, slip_filename TEXT, notes TEXT, created_at TEXT, FOREIGN KEY(event_id) REFERENCES events(id))''')
    c.execute(f'''CREATE TABLE IF NOT EXISTS registration_members (id {id_col}, registration_id INTEGER NOT NULL, member_name TEXT NOT NULL, member_idcard TEXT, idcard_file TEXT, FOREIGN KEY(registration_id) REFERENCES registrations(id))''')
    c.execute(f'''CREATE TABLE IF NOT EXISTS certificates (id {id_col}, registration_id INTEGER NOT NULL, member_id INTEGER, certificate_type TEXT NOT NULL DEFAULT 'individual', verification_code TEXT UNIQUE NOT NULL, issued_at TEXT NOT NULL, FOREIGN KEY(registration_id) REFERENCES registrations(id))''')
    c.execute(f'''CREATE TABLE IF NOT EXISTS import_logs (id {id_col}, tournament_id INTEGER, event_id INTEGER, import_type TEXT, filename TEXT, imported_count INTEGER DEFAULT 0, error_count INTEGER DEFAULT 0, created_at TEXT)''')
    for name, ddl in [
        ('certificates_enabled','INTEGER NOT NULL DEFAULT 0'),('certificate_self_download','INTEGER NOT NULL DEFAULT 1'),
        ('certificate_require_approval','INTEGER NOT NULL DEFAULT 1'),('cert_org','TEXT'),('cert_date','TEXT'),('cert_place','TEXT'),
        ('cert_signer','TEXT'),('cert_signer_position','TEXT'),('cert_style',"TEXT NOT NULL DEFAULT 'navy_gold'"),
        ('cert_heading','TEXT'),('cert_footer_note','TEXT'),('cert_logo_1','TEXT'),('cert_logo_2','TEXT'),('cert_logo_3','TEXT'),
        ('cert_background','TEXT'),('cert_signature','TEXT'),('cert_stamp','TEXT')]: ensure_column(c,'tournaments',name,ddl)
    for name, ddl in [
        ('has_fee','INTEGER NOT NULL DEFAULT 0'),('fee_per','TEXT NOT NULL DEFAULT \'team\''),('require_slip','INTEGER NOT NULL DEFAULT 0'),
        ('has_limit','INTEGER NOT NULL DEFAULT 1'),('waitlist_enabled','INTEGER NOT NULL DEFAULT 0'),('waitlist_limit','INTEGER NOT NULL DEFAULT 0'),
        ('event_mode',"TEXT NOT NULL DEFAULT 'competition'"),('sport_name','TEXT'),('fixed_member_count','INTEGER NOT NULL DEFAULT 0')]: ensure_column(c,'events',name,ddl)
    for name, ddl in [
        ('affiliation','TEXT'),('member_count','INTEGER NOT NULL DEFAULT 1'),('source','TEXT NOT NULL DEFAULT \'web\''),
        ('registration_code','TEXT'),('status','TEXT NOT NULL DEFAULT \'pending\''),('is_waitlist','INTEGER NOT NULL DEFAULT 0'),('approved_at','TEXT'),('award_result',"TEXT NOT NULL DEFAULT 'participant'"),('award_custom','TEXT'),('award_updated_at','TEXT'),('is_complete','INTEGER NOT NULL DEFAULT 1'),('coach_name','TEXT')]: ensure_column(c,'registrations',name,ddl)
    c.execute("UPDATE events SET has_fee = CASE WHEN fee > 0 THEN 1 ELSE has_fee END")
    c.execute("UPDATE events SET has_limit = CASE WHEN max_slots > 0 THEN 1 ELSE 0 END")
    c.execute("UPDATE registrations SET registration_code = 'REG-' || LPAD(CAST(id AS TEXT), 6, '0') WHERE registration_code IS NULL OR registration_code = ''" if IS_POSTGRES else "UPDATE registrations SET registration_code = 'REG-' || printf('%06d', id) WHERE registration_code IS NULL OR registration_code = ''")
    c.execute("UPDATE registrations SET member_count = (SELECT COUNT(*) FROM registration_members m WHERE m.registration_id = registrations.id) WHERE member_count IS NULL OR member_count < 1")
    c.execute("""UPDATE registrations SET is_complete = CASE
        WHEN EXISTS (SELECT 1 FROM events e WHERE e.id = registrations.event_id AND e.event_mode = 'certificate_only') THEN 1
        WHEN (SELECT COUNT(*) FROM registration_members m WHERE m.registration_id = registrations.id) >= member_count THEN 1
        ELSE 0 END""")
    c.execute('CREATE INDEX IF NOT EXISTS idx_events_tournament_id ON events(tournament_id)')
    c.execute('CREATE INDEX IF NOT EXISTS idx_registrations_event_id ON registrations(event_id)')
    c.execute('CREATE INDEX IF NOT EXISTS idx_registration_members_registration_id ON registration_members(registration_id)')
    c.execute('CREATE INDEX IF NOT EXISTS idx_certificates_registration_id ON certificates(registration_id)')
    if not c.execute('SELECT id FROM users WHERE username=?',('admin',)).fetchone():
        c.execute('INSERT INTO users(username,password,role) VALUES(?,?,?)',('admin',generate_password_hash('1234'),'admin'))
    conn.commit(); conn.close()


def is_logged_in(): return 'user_id' in session

def allowed_file(filename, exts=ALLOWED_EXTENSIONS): return '.' in filename and filename.rsplit('.',1)[1].lower() in exts

def now(): return datetime.now().strftime('%Y-%m-%d %H:%M:%S')

def row_value(row, key, default=None):
    try:
        if hasattr(row, 'keys') and key in row.keys():
            return row[key]
        if isinstance(row, dict):
            return row.get(key, default)
    except Exception:
        pass
    return default

def category_label(v): return {'single':'เดี่ยว','pair':'คู่','team':'ทีม'}.get(v,v)

def gender_label(v): return {'male':'ชาย','female':'หญิง','mixed':'ผสม','open':'ไม่ระบุ'}.get(v,v)

def age_label(v): return {'youth':'เยาวชน','general':'ทั่วไป','senior':'อาวุโส'}.get(v,v)

def fee_per_label(v): return {'person':'คน','team':'ทีม'}.get(v,v)

def event_mode_label(v): return {'competition':'สมัครแข่งขันปกติ','certificate_only':'ออกเกียรติบัตรอย่างเดียว'}.get(v or 'competition', v or 'สมัครแข่งขันปกติ')

def is_certificate_only_event(e): return (row_value(e, 'event_mode', 'competition') or 'competition') == 'certificate_only'

def sport_name_text(e):
    raw=str(row_value(e, 'sport_name', '') or '').strip()
    return raw or 'เปตอง'

def sport_competition_label(e):
    raw=sport_name_text(e)
    return raw if raw.startswith('กีฬา') else f'กีฬา{raw}'


def cert_category_line(e):
    """ข้อความประเภท/เพศ/รุ่น สำหรับแสดงบนเกียรติบัตร โดยไม่ซ้ำชื่องานแข่งขัน"""
    category = str(category_label(row_value(e, 'category_type', '')) or '').strip()
    gender = str(gender_label(row_value(e, 'gender_type', '')) or '').strip()
    age_raw = str(row_value(e, 'age_group', '') or '').strip()
    age = str(age_label(age_raw) if age_raw in {'youth', 'general', 'senior'} else age_raw).strip()

    kind = category
    if gender and gender != 'ไม่ระบุ':
        kind = f"{kind}{gender}" if kind else gender

    parts = []
    if kind:
        parts.append(kind)
    if age:
        if age.startswith('รุ่น'):
            parts.append(age)
        elif age.startswith('อายุ'):
            parts.append(f'รุ่น{age}')
        elif age == 'ทั่วไป':
            parts.append('รุ่นทั่วไป')
        else:
            parts.append(f'รุ่น{age}')
    return ' '.join(parts).strip()

def event_member_count_label(e):
    if is_certificate_only_event(e):
        n=int(row_value(e, 'fixed_member_count', 0) or row_value(e, 'team_size', 0) or 0)
        return f'{n} คน' if n > 0 else 'กำหนดเอง'
    if e['category_type']=='single': return '1 คน'
    if e['category_type']=='pair' and e['gender_type']=='mixed': return '2 คน'
    if e['category_type']=='pair': return '2–3 คน'
    return '3–4 คน'

def award_label(v, custom=None):
    labels={
        'participant':'เข้าร่วมการแข่งขัน',
        'champion':'ชนะเลิศ',
        'runner_up_1':'รองชนะเลิศอันดับ 1',
        'runner_up_2':'รองชนะเลิศอันดับ 2',
        'runner_up_3':'รองชนะเลิศอันดับ 3',
        'honorable':'รางวัลชมเชย',
        'custom':(custom or 'รางวัลพิเศษ')
    }
    return labels.get(v or 'participant', v or 'เข้าร่วมการแข่งขัน')

def event_display_name(e):
    custom=(e['event_name'] or '').strip()
    if custom:
        return custom
    base=f"{category_label(e['category_type'])} {gender_label(e['gender_type'])} {age_label(e['age_group'])}"
    if str(row_value(e, 'sport_name', '') or '').strip():
        return f"{sport_competition_label(e)} {base}"
    return base

def allowed_member_counts(category, gender_type=None, fixed_member_count=None):
    try:
        fixed=int(fixed_member_count or 0)
    except (TypeError, ValueError):
        fixed=0
    if fixed > 0:
        return [fixed]
    if category=='single':
        return [1]
    if category=='pair':
        return [2] if gender_type=='mixed' else [2,3]
    return [3,4]

def suggested_member_count(category, gender_type=None, fixed_member_count=None): return max(allowed_member_counts(category, gender_type, fixed_member_count))

def event_allowed_member_counts(e):
    return allowed_member_counts(e['category_type'], e['gender_type'], row_value(e, 'fixed_member_count', 0) if is_certificate_only_event(e) else None)

def event_suggested_member_count(e):
    return max(event_allowed_member_counts(e))

def registration_is_complete(member_count, members):
    try: expected=int(member_count or 0)
    except (TypeError,ValueError): expected=0
    named=[]
    for m in members:
        name=m.get('name') if isinstance(m,dict) else m[0]
        if str(name or '').strip(): named.append(name)
    return 1 if expected > 0 and len(named) >= expected else 0

def incomplete_import_allowed(event, team_name, members):
    # เดี่ยวต้องมีชื่อนักกีฬาเสมอ ส่วนคู่/ทีมสามารถส่งรายชื่อมาไม่ครบแล้วแก้ภายหลังได้
    if event['category_type']=='single': return len(members)==1
    return bool(team_name or members)

def _zero_event_count():
    return {'total': 0, 'active': 0, 'waitlist': 0, 'approved': 0, 'pending': 0, 'incomplete': 0, 'awarded': 0}


def event_counts_map(conn, event_ids=None, tournament_id=None):
    """นับจำนวนผู้สมัครแบบรวมครั้งเดียว ลดปัญหาเปิด DB ซ้ำหลายรอบจนระบบช้า"""
    params = []
    where = ''
    if event_ids is not None:
        ids = [int(x) for x in event_ids if x is not None]
        if not ids:
            return {}
        marks = ','.join(['?'] * len(ids))
        where = f'WHERE e.id IN ({marks})'
        params = ids
    elif tournament_id is not None:
        where = 'WHERE e.tournament_id=?'
        params = [tournament_id]
    else:
        return {}
    rows = conn.execute(f'''
        SELECT e.id event_id,
            COUNT(r.id) total,
            SUM(CASE WHEN r.id IS NOT NULL AND COALESCE(r.is_waitlist,0)=0 THEN 1 ELSE 0 END) active,
            SUM(CASE WHEN r.id IS NOT NULL AND COALESCE(r.is_waitlist,0)=1 THEN 1 ELSE 0 END) waitlist,
            SUM(CASE WHEN r.id IS NOT NULL AND r.status='approved' THEN 1 ELSE 0 END) approved,
            SUM(CASE WHEN r.id IS NOT NULL AND r.status!='approved' THEN 1 ELSE 0 END) pending,
            SUM(CASE WHEN r.id IS NOT NULL AND COALESCE(r.is_complete,0)=0 THEN 1 ELSE 0 END) incomplete,
            SUM(CASE WHEN r.id IS NOT NULL AND r.award_result IS NOT NULL AND r.award_result!='participant' THEN 1 ELSE 0 END) awarded
        FROM events e
        LEFT JOIN registrations r ON r.event_id=e.id
        {where}
        GROUP BY e.id
    ''', params).fetchall()
    out = {}
    for row in rows:
        out[int(row['event_id'])] = {
            'total': int(row['total'] or 0),
            'active': int(row['active'] or 0),
            'waitlist': int(row['waitlist'] or 0),
            'approved': int(row['approved'] or 0),
            'pending': int(row['pending'] or 0),
            'incomplete': int(row['incomplete'] or 0),
            'awarded': int(row['awarded'] or 0),
        }
    return out


def event_reg_count(event_id, include_waitlist=False, conn=None):
    close_conn = False
    if conn is None:
        conn = get_db(); close_conn = True
    q='SELECT COUNT(*) total FROM registrations WHERE event_id=?'
    args=[event_id]
    if not include_waitlist: q += ' AND COALESCE(is_waitlist,0)=0'
    total=conn.execute(q,args).fetchone()['total']
    if close_conn: conn.close()
    return total


def registration_members_map(conn, registration_ids, select_cols='id, registration_id, member_name, member_idcard, idcard_file'):
    ids = [int(x) for x in registration_ids if x is not None]
    if not ids:
        return {}
    marks = ','.join(['?'] * len(ids))
    rows = conn.execute(f'SELECT {select_cols} FROM registration_members WHERE registration_id IN ({marks}) ORDER BY registration_id,id', ids).fetchall()
    out = {rid: [] for rid in ids}
    for row in rows:
        out.setdefault(int(row['registration_id']), []).append(row)
    return out

def save_uploaded_file(file_obj,prefix='file'):
    if not file_obj or not file_obj.filename: return None
    if not allowed_file(file_obj.filename): return None
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
    ext=secure_filename(file_obj.filename).rsplit('.',1)[1].lower()
    name=f"{prefix}_{datetime.now().strftime('%Y%m%d%H%M%S%f')}.{ext}"
    file_obj.save(os.path.join(UPLOAD_FOLDER,name)); return name

def delete_uploaded_file(name):
    if name:
        p=os.path.join(UPLOAD_FOLDER,name)
        if os.path.exists(p): os.remove(p)

def save_certificate_asset(file_obj, prefix):
    if not file_obj or not file_obj.filename:
        return None
    if not allowed_file(file_obj.filename, CERT_IMAGE_EXTENSIONS):
        return None
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
    ext=secure_filename(file_obj.filename).rsplit('.',1)[1].lower()
    name=f"{prefix}_{datetime.now().strftime('%Y%m%d%H%M%S%f')}.{ext}"
    file_obj.save(os.path.join(UPLOAD_FOLDER,name))
    return name

def truthy(value): return str(value or '').strip().lower() in {'1','true','yes','y','ใช่','มี','เปิด','on'}

def normalize_category(v): return {'เดี่ยว':'single','single':'single','คู่':'pair','pair':'pair','ทีม':'team','team':'team'}.get(str(v or '').strip().lower(), '')

def normalize_gender(v): return {'ชาย':'male','male':'male','หญิง':'female','female':'female','ผสม':'mixed','mixed':'mixed','ไม่ระบุ':'open','open':'open'}.get(str(v or '').strip().lower(), 'open')

def normalize_age(v):
    raw=str(v or '').strip(); return {'เยาวชน':'youth','youth':'youth','ทั่วไป':'general','general':'general','อาวุโส':'senior','senior':'senior'}.get(raw.lower(), raw or 'general')

def unique_registration_code(conn):
    while True:
        code='REG-'+datetime.now().strftime('%y%m')+'-'+secrets.token_hex(3).upper()
        if not conn.execute('SELECT id FROM registrations WHERE registration_code=?',(code,)).fetchone(): return code

def registration_capacity_state(event):
    active=event_reg_count(event['id'])
    if not event['has_limit'] or int(event['max_slots'] or 0)<=0: return False, False
    full=active >= int(event['max_slots'])
    if not full or not event['waitlist_enabled']: return full, False
    wait_total=event_reg_count(event['id'],True)-active
    wait_limit=int(event['waitlist_limit'] or 0)
    return True, bool(wait_limit<=0 or wait_total<wait_limit)

def get_owned_tournament(conn,tournament_id):
    return conn.execute('SELECT * FROM tournaments WHERE id=? AND created_by=?',(tournament_id,session.get('user_id'))).fetchone()

def get_owned_event(conn,event_id):
    return conn.execute('''SELECT e.*,t.title tournament_title,t.created_by,t.certificates_enabled,t.certificate_self_download,t.certificate_require_approval,t.cert_org,t.cert_date,t.cert_place,t.cert_signer,t.cert_signer_position,t.cert_style,t.cert_heading,t.cert_footer_note,t.cert_logo_1,t.cert_logo_2,t.cert_logo_3,t.cert_background,t.cert_signature,t.cert_stamp FROM events e JOIN tournaments t ON e.tournament_id=t.id WHERE e.id=?''',(event_id,)).fetchone()

@app.context_processor
def helpers(): return dict(category_label=category_label,gender_label=gender_label,age_label=age_label,fee_per_label=fee_per_label,award_label=award_label,event_display_name=event_display_name,allowed_member_counts=allowed_member_counts,event_allowed_member_counts=event_allowed_member_counts,event_suggested_member_count=event_suggested_member_count,event_mode_label=event_mode_label,event_member_count_label=event_member_count_label,sport_competition_label=sport_competition_label,cert_category_line=cert_category_line,cert_issue_date=format_cert_issue_date,is_certificate_only_event=is_certificate_only_event)

@app.route('/uploads/<path:filename>')
def uploaded_file(filename): return send_from_directory(UPLOAD_FOLDER,filename)

@app.route('/')
def home():
    conn=get_db()
    tournaments=conn.execute('SELECT * FROM tournaments ORDER BY id DESC').fetchall()
    event_map={}; count_map={}
    tournament_ids=[t['id'] for t in tournaments]
    if tournament_ids:
        marks=','.join(['?']*len(tournament_ids))
        events=conn.execute(f'SELECT * FROM events WHERE tournament_id IN ({marks}) ORDER BY tournament_id,id',tournament_ids).fetchall()
        counts=event_counts_map(conn,event_ids=[e['id'] for e in events])
        for t in tournaments:
            event_map[t['id']]=[]; count_map[t['id']]={}
        for e in events:
            event_map[e['tournament_id']].append(e)
            count_map[e['tournament_id']][e['id']]=counts.get(e['id'],_zero_event_count())['active']
    conn.close()
    return render_template('home.html',tournaments=tournaments,event_map=event_map,count_map=count_map)


@app.route('/tournament/<int:tournament_id>/registrations')
def public_tournament_registrations(tournament_id):
    """Public applicant list. Intentionally excludes contact details, ID cards and private tracking codes."""
    conn=get_db()
    tournament=conn.execute('SELECT * FROM tournaments WHERE id=?',(tournament_id,)).fetchone()
    if not tournament:
        conn.close(); abort(404)
    events=conn.execute('SELECT * FROM events WHERE tournament_id=? ORDER BY age_group,category_type,gender_type,id',(tournament_id,)).fetchall()
    selected_event_id=request.args.get('event_id',type=int)
    selected_event=None
    if selected_event_id:
        selected_event=next((e for e in events if e['id']==selected_event_id),None)
        if not selected_event: selected_event_id=None
    query='SELECT r.id,r.event_id,r.registration_code,r.team_name,r.affiliation,r.member_count,r.is_complete,r.is_waitlist,r.created_at,e.event_name,e.category_type,e.gender_type,e.age_group,e.event_mode,e.sport_name,e.fixed_member_count FROM registrations r JOIN events e ON r.event_id=e.id WHERE e.tournament_id=?'
    args=[tournament_id]
    if selected_event_id:
        query += ' AND e.id=?'; args.append(selected_event_id)
    query += ' ORDER BY e.id,r.id'
    rows=conn.execute(query,args).fetchall()
    members_map=registration_members_map(conn,[r['id'] for r in rows],select_cols='id, registration_id, member_name')
    count_rows=event_counts_map(conn,event_ids=[e['id'] for e in events])
    count_map={e['id']:count_rows.get(e['id'],_zero_event_count())['active'] for e in events}
    wait_map={e['id']:count_rows.get(e['id'],_zero_event_count())['waitlist'] for e in events}
    conn.close()
    return render_template('public_registrations.html',tournament=tournament,events=events,selected_event=selected_event,selected_event_id=selected_event_id,rows=rows,members_map=members_map,count_map=count_map,wait_map=wait_map)

@app.route('/event/<int:event_id>/register',methods=['GET','POST'])
def register_event(event_id):
    conn=get_db(); event=conn.execute('SELECT * FROM events WHERE id=?',(event_id,)).fetchone()
    if not event: conn.close(); flash('ไม่พบอีเวนต์'); return redirect(url_for('home'))
    tournament=conn.execute('SELECT * FROM tournaments WHERE id=?',(event['tournament_id'],)).fetchone(); conn.close()
    cert_only=is_certificate_only_event(event)
    fixed_count=int(row_value(event,'fixed_member_count',0) or row_value(event,'team_size',0) or 0) if cert_only else 0
    allowed_counts=allowed_member_counts(event['category_type'], event['gender_type'], fixed_count if cert_only else None)
    default_member_count=suggested_member_count(event['category_type'], event['gender_type'], fixed_count if cert_only else None)
    reg_count=event_reg_count(event_id); full, can_waitlist=registration_capacity_state(event)
    if request.method=='POST':
        if not event['is_open'] or not tournament['is_open']: flash('อีเวนต์นี้ปิดรับสมัครแล้ว'); return redirect(url_for('register_event',event_id=event_id))
        if full and not can_waitlist: flash('อีเวนต์นี้เต็มแล้ว'); return redirect(url_for('register_event',event_id=event_id))
        member_count=parse_int(request.form.get('member_count',default_member_count),default_member_count,1,99)
        if member_count not in allowed_counts: flash('จำนวนผู้เล่นไม่ตรงตามที่แอดมินกำหนด'); return redirect(url_for('register_event',event_id=event_id))
        team_name=request.form.get('team_name','').strip()
        affiliation=request.form.get('affiliation','').strip()
        coach_name=request.form.get('coach_name','').strip()
        contact=request.form.get('contact_name','').strip()
        phone=request.form.get('phone','').strip()
        notes=request.form.get('notes','').strip()
        if cert_only:
            if not team_name: flash('กรุณากรอกชื่อทีม'); return redirect(url_for('register_event',event_id=event_id))
            if not coach_name: flash('กรุณากรอกชื่อผู้ฝึกสอน'); return redirect(url_for('register_event',event_id=event_id))
            contact=contact or coach_name
        else:
            if not contact or not phone: flash('กรุณากรอกชื่อผู้ติดต่อและเบอร์โทร'); return redirect(url_for('register_event',event_id=event_id))
            if event['category_type']=='team' and not team_name: flash('ประเภททีมต้องกรอกชื่อทีม'); return redirect(url_for('register_event',event_id=event_id))
        members=[]
        for i in range(1,member_count+1):
            n=request.form.get(f'member_name_{i}','').strip()
            idc='' if cert_only else request.form.get(f'member_idcard_{i}','').strip()
            f=None if cert_only else request.files.get(f'idcard_file_{i}')
            fn=None
            # โหมดออกเกียรติบัตรอย่างเดียว: เว้นว่างได้ ช่องว่าง = ไม่มีนักกีฬา ไม่ออกใบรายบุคคล
            # โหมดสมัครแข่งขันปกติ: ยังบังคับกรอกชื่อนักกีฬาตามจำนวนเดิม
            if not n:
                if cert_only:
                    continue
                flash(f'กรุณากรอกชื่อนักกีฬาคนที่ {i}')
                return redirect(url_for('register_event',event_id=event_id))
            if f and f.filename:
                fn=save_uploaded_file(f,f'idcard_{i}')
                if not fn: flash('ไฟล์บัตรประชาชนต้องเป็น JPG PNG WEBP หรือ PDF'); return redirect(url_for('register_event',event_id=event_id))
            members.append((n,idc,fn))
        slip=None; sf=None if cert_only else request.files.get('slip_file')
        if sf and sf.filename:
            slip=save_uploaded_file(sf,'slip')
            if not slip: flash('ไฟล์สลิปต้องเป็น JPG PNG WEBP หรือ PDF'); return redirect(url_for('register_event',event_id=event_id))
        if (not cert_only) and event['has_fee'] and event['require_slip'] and not slip: flash('กรุณาแนบหลักฐานการชำระเงิน'); return redirect(url_for('register_event',event_id=event_id))
        conn=get_db(); code=unique_registration_code(conn); wait=1 if full and can_waitlist else 0
        status='approved' if cert_only else 'pending'
        source='certificate_only' if cert_only else 'web'
        complete=1 if cert_only else registration_is_complete(member_count,[{'name':m[0]} for m in members])
        rid=insert_returning_id(conn,'''INSERT INTO registrations(event_id,team_name,affiliation,contact_name,phone,slip_filename,notes,created_at,member_count,source,registration_code,status,is_waitlist,is_complete,coach_name,approved_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',(event_id,team_name or None,affiliation or None,contact,phone,slip,notes,now(),member_count,source,code,status,wait,complete,coach_name or None,now() if cert_only else None)); c=conn.cursor()
        for m in members: c.execute('INSERT INTO registration_members(registration_id,member_name,member_idcard,idcard_file) VALUES(?,?,?,?)',(rid,*m))
        conn.commit(); conn.close(); return redirect(url_for('registration_status',code=code))
    return render_template('register_event.html',event=event,tournament=tournament,reg_count=reg_count,full=full,can_waitlist=can_waitlist,default_member_count=default_member_count,allowed_counts=allowed_counts,cert_only=cert_only)

@app.route('/registration/<code>')
def registration_status(code):
    conn=get_db(); reg=conn.execute('''SELECT r.*,e.event_name,e.category_type,e.gender_type,e.age_group,e.has_fee,e.fee,e.fee_per,e.event_mode,e.sport_name,e.fixed_member_count,t.title tournament_title,t.certificates_enabled,t.certificate_self_download,t.certificate_require_approval FROM registrations r JOIN events e ON r.event_id=e.id JOIN tournaments t ON e.tournament_id=t.id WHERE r.registration_code=?''',(code,)).fetchone()
    if not reg: conn.close(); abort(404)
    members=conn.execute('SELECT * FROM registration_members WHERE registration_id=? ORDER BY id',(reg['id'],)).fetchall(); conn.close()
    instant=is_certificate_only_event(reg)
    cert_ready=bool((instant or reg['is_complete']) and reg['certificates_enabled'] and reg['certificate_self_download'] and (instant or not reg['certificate_require_approval'] or reg['status']=='approved'))
    return render_template('registration_status.html',reg=reg,members=members,cert_ready=cert_ready,cert_only=instant)

@app.route('/registration/<code>/edit', methods=['GET','POST'])
def public_edit_registration(code):
    """ให้ผู้สมัครแก้ไขชื่อทีมและรายชื่อนักกีฬาในช่องที่กำหนดไว้ได้เอง
    - ไม่มีปุ่มลบรายการสมัคร
    - ช่องชื่อนักกีฬาที่เว้นว่าง = ไม่มีนักกีฬา และจะไม่ออกเกียรติบัตรรายบุคคล
    - เติมชื่อกลับเข้าไปในช่องว่างได้ภายหลัง
    """
    conn=get_db()
    reg=conn.execute('''SELECT r.*,e.event_name,e.category_type,e.gender_type,e.age_group,e.has_fee,e.fee,e.fee_per,e.event_mode,e.sport_name,e.fixed_member_count,t.title tournament_title,t.certificates_enabled,t.certificate_self_download,t.certificate_require_approval
                        FROM registrations r
                        JOIN events e ON r.event_id=e.id
                        JOIN tournaments t ON e.tournament_id=t.id
                        WHERE r.registration_code=?''',(code,)).fetchone()
    if not reg:
        conn.close(); abort(404)
    members=conn.execute('SELECT * FROM registration_members WHERE registration_id=? ORDER BY id',(reg['id'],)).fetchall()
    slot_count=max(int(reg['member_count'] or 0), len(members), 1)

    if request.method=='POST':
        team_name=request.form.get('team_name','').strip()
        team_required=is_certificate_only_event(reg) or reg['category_type']=='team'
        if team_required and not team_name:
            conn.close(); flash('กรุณากรอกชื่อทีม')
            return redirect(url_for('public_edit_registration',code=code))

        new_members=[]
        removed_files=[]
        for i in range(1,slot_count+1):
            name=request.form.get(f'member_name_{i}','').strip()
            old_member=members[i-1] if i-1 < len(members) else None
            if name:
                new_members.append((name, old_member['member_idcard'] if old_member else '', old_member['idcard_file'] if old_member else None))
            elif old_member and old_member['idcard_file']:
                removed_files.append(old_member['idcard_file'])

        # เก็บจำนวนช่องไว้เท่าเดิม เพื่อให้ผู้สมัครกลับมาเติมชื่อในช่องว่างได้
        complete=1 if is_certificate_only_event(reg) else registration_is_complete(slot_count,[{'name':m[0]} for m in new_members])
        status=reg['status']
        if status=='approved' and not complete:
            status='pending'

        conn.execute('DELETE FROM certificates WHERE registration_id=?',(reg['id'],))
        conn.execute('DELETE FROM registration_members WHERE registration_id=?',(reg['id'],))
        conn.execute('UPDATE registrations SET team_name=?, is_complete=?, status=? WHERE id=?',(team_name or None,complete,status,reg['id']))
        for file_name in removed_files:
            delete_uploaded_file(file_name)
        for m in new_members:
            conn.execute('INSERT INTO registration_members(registration_id,member_name,member_idcard,idcard_file) VALUES(?,?,?,?)',(reg['id'],*m))
        conn.commit(); conn.close()
        flash('แก้ไขข้อมูลทีมเรียบร้อยแล้ว ช่องว่างจะไม่นับเป็นนักกีฬาและไม่ออกเกียรติบัตรรายบุคคล')
        return redirect(url_for('registration_status',code=code))

    slot_members=[]
    for i in range(slot_count):
        slot_members.append(members[i] if i < len(members) else None)
    conn.close()
    return render_template('public_edit_registration.html',reg=reg,members=members,slot_members=slot_members,slot_count=slot_count)

@app.route('/certificate-search')
def certificate_search():
    """Public self-service certificate search by athlete or coach name."""
    query=(request.args.get('q') or '').strip()
    results=[]
    searched=False
    if query:
        searched=True
        if len(query) < 2:
            flash('กรุณากรอกชื่ออย่างน้อย 2 ตัวอักษร')
        else:
            conn=get_db()
            like=f'%{query}%'
            results=conn.execute('''
                SELECT m.id member_id,m.member_name,'athlete' AS person_type,r.id registration_id,r.team_name,r.affiliation,r.coach_name,r.status,r.is_waitlist,r.is_complete,r.award_result,r.award_custom,
                e.event_name,e.category_type,e.gender_type,e.age_group,e.event_mode,e.sport_name,e.fixed_member_count,t.title tournament_title,t.certificates_enabled,t.certificate_self_download,t.certificate_require_approval
                FROM registration_members m
                JOIN registrations r ON m.registration_id=r.id
                JOIN events e ON r.event_id=e.id
                JOIN tournaments t ON e.tournament_id=t.id
                WHERE m.member_name LIKE ?
                UNION ALL
                SELECT NULL AS member_id,r.coach_name AS member_name,'coach' AS person_type,r.id registration_id,r.team_name,r.affiliation,r.coach_name,r.status,r.is_waitlist,r.is_complete,r.award_result,r.award_custom,
                e.event_name,e.category_type,e.gender_type,e.age_group,e.event_mode,e.sport_name,e.fixed_member_count,t.title tournament_title,t.certificates_enabled,t.certificate_self_download,t.certificate_require_approval
                FROM registrations r
                JOIN events e ON r.event_id=e.id
                JOIN tournaments t ON e.tournament_id=t.id
                WHERE r.coach_name IS NOT NULL AND r.coach_name <> '' AND r.coach_name LIKE ?
                ORDER BY tournament_title DESC,event_name,registration_id,person_type
            ''',(like,like)).fetchall()
            conn.close()
    return render_template('certificate_search.html',query=query,results=results,searched=searched)

@app.route('/login',methods=['GET','POST'])
def login():
    if request.method=='POST':
        conn=get_db(); user=conn.execute('SELECT * FROM users WHERE username=?',(request.form.get('username','').strip(),)).fetchone(); conn.close()
        if user and check_password_hash(user['password'],request.form.get('password','')):
            session.update(user_id=user['id'],username=user['username'],role=user['role']); flash('เข้าสู่ระบบสำเร็จ'); return redirect(url_for('admin_dashboard'))
        flash('ชื่อผู้ใช้หรือรหัสผ่านไม่ถูกต้อง')
    return render_template('login.html')
@app.route('/logout')
def logout(): session.clear(); flash('ออกจากระบบแล้ว'); return redirect(url_for('home'))

@app.route('/admin')
def admin_dashboard():
    if not is_logged_in(): return redirect(url_for('login'))
    conn=get_db()
    tournaments=conn.execute('SELECT * FROM tournaments WHERE created_by=? ORDER BY id DESC',(session['user_id'],)).fetchall()
    event_map={}; count_map={}; dashboard={'tournaments':len(tournaments),'events':0,'registrations':0,'open_events':0}
    tournament_ids=[t['id'] for t in tournaments]
    if tournament_ids:
        marks=','.join(['?']*len(tournament_ids))
        events=conn.execute(f'SELECT * FROM events WHERE tournament_id IN ({marks}) ORDER BY tournament_id,id',tournament_ids).fetchall()
        counts=event_counts_map(conn,event_ids=[e['id'] for e in events])
        for t in tournaments:
            event_map[t['id']]=[]; count_map[t['id']]={}
        for e in events:
            n=counts.get(e['id'],_zero_event_count())['active']
            event_map[e['tournament_id']].append(e)
            count_map[e['tournament_id']][e['id']]=n
            dashboard['events'] += 1
            dashboard['registrations'] += n
            dashboard['open_events'] += 1 if e['is_open'] else 0
    conn.close()
    return render_template('admin_dashboard.html',tournaments=tournaments,event_map=event_map,count_map=count_map,dashboard=dashboard)

@app.route('/admin/tournament/create',methods=['GET','POST'])
def create_tournament():
    if not is_logged_in(): return redirect(url_for('login'))
    if request.method=='POST':
        title=request.form.get('title','').strip(); desc=request.form.get('description','').strip()
        if not title: flash('กรุณากรอกชื่องานแข่งขัน'); return redirect(url_for('create_tournament'))
        conn=get_db(); tid=insert_returning_id(conn,'INSERT INTO tournaments(title,description,created_by,created_at,is_open) VALUES(?,?,?,?,?)',(title,desc,session['user_id'],now(),1 if request.form.get('is_open') else 0)); conn.commit(); conn.close(); return redirect(url_for('manage_events',tournament_id=tid))
    return render_template('create_tournament.html')

@app.route('/admin/tournament/<int:tournament_id>/edit',methods=['GET','POST'])
def edit_tournament(tournament_id):
    if not is_logged_in(): return redirect(url_for('login'))
    conn=get_db(); t=get_owned_tournament(conn,tournament_id)
    if not t: conn.close(); flash('ไม่พบงานแข่งขัน'); return redirect(url_for('admin_dashboard'))
    if request.method=='POST':
        title=request.form.get('title','').strip()
        if not title: conn.close(); flash('กรุณากรอกชื่องานแข่งขัน'); return redirect(url_for('edit_tournament',tournament_id=tournament_id))
        conn.execute('UPDATE tournaments SET title=?,description=?,is_open=? WHERE id=?',(title,request.form.get('description','').strip(),1 if request.form.get('is_open') else 0,tournament_id)); conn.commit(); conn.close(); return redirect(url_for('admin_dashboard'))
    conn.close(); return render_template('edit_tournament.html',tournament=t)

@app.route('/admin/tournament/<int:tournament_id>/events')
def manage_events(tournament_id):
    if not is_logged_in(): return redirect(url_for('login'))
    conn=get_db(); t=get_owned_tournament(conn,tournament_id)
    if not t: conn.close(); return redirect(url_for('admin_dashboard'))
    events=conn.execute('SELECT * FROM events WHERE tournament_id=? ORDER BY id',(tournament_id,)).fetchall()
    count_rows=event_counts_map(conn,event_ids=[e['id'] for e in events])
    counts={e['id']:count_rows.get(e['id'],_zero_event_count())['active'] for e in events}
    waitcounts={e['id']:count_rows.get(e['id'],_zero_event_count())['waitlist'] for e in events}
    total=sum(counts.values())
    conn.close()
    return render_template('manage_events.html',tournament=t,events=events,counts=counts,waitcounts=waitcounts,total_regs=total)


def parse_int(value, default=0, min_value=None, max_value=None):
    try:
        n=int(value)
    except (TypeError, ValueError):
        n=default
    if min_value is not None: n=max(min_value,n)
    if max_value is not None: n=min(max_value,n)
    return n

def parse_event_form():
    event_mode=request.form.get('event_mode','competition')
    if event_mode not in {'competition','certificate_only'}:
        event_mode='competition'
    category=request.form.get('category_type','single')
    gender=request.form.get('gender_type','open')
    sport_name=request.form.get('sport_name','').strip()
    fixed_member_count=parse_int(request.form.get('fixed_member_count','0'),0,0,99) if event_mode=='certificate_only' else 0
    has_fee=1 if request.form.get('has_fee') else 0
    has_limit=1 if request.form.get('has_limit') else 0
    fee=parse_int(request.form.get('fee','0'),0,0) if has_fee else 0
    max_slots=parse_int(request.form.get('max_slots','0'),0,0) if has_limit else 0
    team_size=fixed_member_count if fixed_member_count > 0 else suggested_member_count(category, gender)
    return dict(
        event_name=request.form.get('event_name','').strip(),
        category_type=category,
        gender_type=gender,
        age_group=request.form.get('age_group','general').strip(),
        team_size=team_size,
        has_fee=has_fee,
        fee=fee,
        fee_per=request.form.get('fee_per','team'),
        require_slip=1 if has_fee and request.form.get('require_slip') else 0,
        has_limit=has_limit,
        max_slots=max_slots,
        waitlist_enabled=1 if has_limit and request.form.get('waitlist_enabled') else 0,
        waitlist_limit=parse_int(request.form.get('waitlist_limit','0'),0,0),
        is_open=1 if request.form.get('is_open') else 0,
        event_mode=event_mode,
        sport_name=sport_name,
        fixed_member_count=fixed_member_count
    )

@app.route('/admin/tournament/<int:tournament_id>/event/create',methods=['GET','POST'])
def create_event(tournament_id):
    if not is_logged_in(): return redirect(url_for('login'))
    conn=get_db(); t=get_owned_tournament(conn,tournament_id)
    if not t: conn.close(); return redirect(url_for('admin_dashboard'))
    if request.method=='POST':
        d=parse_event_form(); cols=','.join(d.keys()); marks=','.join(['?']*len(d)); conn.execute(f'INSERT INTO events(tournament_id,{cols},created_at) VALUES(?,{marks},?)',(tournament_id,*d.values(),now())); conn.commit(); conn.close(); flash('เพิ่มอีเวนต์เรียบร้อยแล้ว'); return redirect(url_for('manage_events',tournament_id=tournament_id))
    conn.close(); return render_template('create_event.html',tournament=t)

@app.route('/admin/event/<int:event_id>/edit',methods=['GET','POST'])
def edit_event(event_id):
    if not is_logged_in(): return redirect(url_for('login'))
    conn=get_db(); e=get_owned_event(conn,event_id)
    if not e or e['created_by']!=session['user_id']: conn.close(); return redirect(url_for('admin_dashboard'))
    t=conn.execute('SELECT * FROM tournaments WHERE id=?',(e['tournament_id'],)).fetchone()
    if request.method=='POST':
        d=parse_event_form(); sets=','.join([f'{k}=?' for k in d]); conn.execute(f'UPDATE events SET {sets} WHERE id=?',(*d.values(),event_id)); conn.commit(); conn.close(); flash('แก้ไขอีเวนต์เรียบร้อยแล้ว'); return redirect(url_for('manage_events',tournament_id=e['tournament_id']))
    conn.close(); return render_template('edit_event.html',tournament=t,event=e)

@app.route('/admin/tournament/<int:tournament_id>/registrations')
def tournament_registrations(tournament_id):
    if not is_logged_in(): return redirect(url_for('login'))
    conn=get_db(); t=get_owned_tournament(conn,tournament_id)
    if not t: conn.close(); return redirect(url_for('admin_dashboard'))
    events=conn.execute('SELECT * FROM events WHERE tournament_id=? ORDER BY age_group,category_type,gender_type,id',(tournament_id,)).fetchall()
    selected_event_id=request.args.get('event_id',type=int)
    selected_event=None
    if selected_event_id:
        selected_event=next((e for e in events if e['id']==selected_event_id),None)
        if not selected_event: selected_event_id=None
    status_filter=request.args.get('status_filter','pending')
    if status_filter not in {'pending','approved'}: status_filter='pending'
    base_query='''SELECT r.*,e.event_name,e.category_type,e.gender_type,e.age_group,e.event_mode,e.sport_name,e.fixed_member_count,e.fee,e.has_fee,e.fee_per FROM registrations r JOIN events e ON r.event_id=e.id WHERE e.tournament_id=?'''
    base_args=[tournament_id]
    if selected_event_id:
        base_query += ' AND e.id=?'; base_args.append(selected_event_id)
    all_rows=conn.execute(base_query + ' ORDER BY e.id,r.id DESC',base_args).fetchall()
    query=base_query + (" AND r.status='approved'" if status_filter=='approved' else " AND r.status!='approved'") + ' ORDER BY e.id,r.id DESC'
    rows=conn.execute(query,base_args).fetchall()
    members_map=registration_members_map(conn,[r['id'] for r in rows])
    count_rows=event_counts_map(conn,event_ids=[e['id'] for e in events])
    summary=[]
    for e in events:
        st=count_rows.get(e['id'],_zero_event_count())
        summary.append(dict(event=e,total=st['total'],approved=st['approved'],pending=st['pending'],waitlist=st['waitlist'],awarded=st['awarded']))
    stats={
        'total':len(all_rows),
        'approved':sum(1 for r in all_rows if r['status']=='approved'),
        'pending':sum(1 for r in all_rows if r['status']!='approved'),
        'waitlist':sum(1 for r in all_rows if r['is_waitlist']),
        'incomplete':sum(1 for r in all_rows if not r['is_complete']),
        'awarded':sum(1 for r in all_rows if (r['award_result'] or 'participant')!='participant'),
        'visible':len(rows)
    }
    conn.close()
    return render_template('tournament_registrations.html',tournament=t,events=events,selected_event=selected_event,selected_event_id=selected_event_id,status_filter=status_filter,rows=rows,members_map=members_map,summary=summary,stats=stats)

@app.route('/admin/registration/<int:registration_id>/approve')
def approve_registration(registration_id):
    if not is_logged_in(): return redirect(url_for('login'))
    conn=get_db(); r=conn.execute('''SELECT r.*,e.tournament_id,t.created_by FROM registrations r JOIN events e ON r.event_id=e.id JOIN tournaments t ON e.tournament_id=t.id WHERE r.id=?''',(registration_id,)).fetchone()
    if not r or r['created_by']!=session['user_id']: conn.close(); return redirect(url_for('admin_dashboard'))
    if r['status']!='approved' and not r['is_complete']:
        conn.close(); flash('ยังอนุมัติไม่ได้ เพราะรายชื่อนักกีฬายังไม่ครบ กรุณากดแก้ไขข้อมูลก่อน')
        return redirect(url_for('tournament_registrations',tournament_id=r['tournament_id'],event_id=r['event_id']))
    new='pending' if r['status']=='approved' else 'approved'; conn.execute('UPDATE registrations SET status=?,approved_at=? WHERE id=?',(new,now() if new=='approved' else None,registration_id)); conn.commit(); conn.close(); flash('อัปเดตสถานะเรียบร้อยแล้ว')
    return_tab=request.args.get('status_filter') or ('approved' if new=='pending' else 'pending')
    return redirect(url_for('tournament_registrations',tournament_id=r['tournament_id'],event_id=r['event_id'],status_filter=return_tab))

@app.route('/admin/registration/<int:registration_id>/edit',methods=['GET','POST'])
def edit_registration(registration_id):
    if not is_logged_in(): return redirect(url_for('login'))
    conn=get_db()
    reg=conn.execute('''SELECT r.*,e.tournament_id,e.event_mode,e.sport_name,e.fixed_member_count,t.created_by FROM registrations r JOIN events e ON r.event_id=e.id JOIN tournaments t ON e.tournament_id=t.id WHERE r.id=?''',(registration_id,)).fetchone()
    if not reg or reg['created_by']!=session['user_id']:
        conn.close(); return redirect(url_for('admin_dashboard'))
    events=conn.execute('SELECT * FROM events WHERE tournament_id=? ORDER BY age_group,category_type,gender_type,id',(reg['tournament_id'],)).fetchall()
    members=conn.execute('SELECT * FROM registration_members WHERE registration_id=? ORDER BY id',(registration_id,)).fetchall()
    if request.method=='POST':
        event_id=request.form.get('event_id',type=int)
        event=next((e for e in events if e['id']==event_id),None)
        if not event:
            conn.close(); flash('ไม่พบอีเวนต์ที่เลือก'); return redirect(request.url)
        try: member_count=int(request.form.get('member_count','0'))
        except ValueError: member_count=0
        if member_count not in allowed_member_counts(event['category_type'],event['gender_type'], event['fixed_member_count'] if is_certificate_only_event(event) else None):
            conn.close(); flash('จำนวนผู้เล่นไม่ตรงตามประเภทการแข่งขัน'); return redirect(request.url)
        team_name=request.form.get('team_name','').strip(); affiliation=request.form.get('affiliation','').strip(); coach_name=request.form.get('coach_name','').strip()
        contact=request.form.get('contact_name','').strip(); phone=request.form.get('phone','').strip(); notes=request.form.get('notes','').strip()
        if is_certificate_only_event(event):
            if not team_name:
                conn.close(); flash('กรุณากรอกชื่อทีม'); return redirect(request.url)
            if not coach_name:
                conn.close(); flash('กรุณากรอกชื่อผู้ฝึกสอน'); return redirect(request.url)
            contact=contact or coach_name
        else:
            if event['category_type']=='team' and not team_name:
                conn.close(); flash('ประเภททีมต้องกรอกชื่อทีม'); return redirect(request.url)
            if not contact or not phone:
                conn.close(); flash('กรุณากรอกชื่อผู้ติดต่อและเบอร์โทร'); return redirect(request.url)
        old_files={m['id']:m['idcard_file'] for m in members if m['idcard_file']}
        kept_files=set(); updated_members=[]
        for i in range(1,member_count+1):
            name=request.form.get(f'member_name_{i}','').strip(); idcard=request.form.get(f'member_idcard_{i}','').strip()
            old_member_id=request.form.get(f'old_member_id_{i}',type=int)
            old_file=old_files.get(old_member_id)
            uploaded=request.files.get(f'idcard_file_{i}')
            file_name=old_file
            if uploaded and uploaded.filename:
                file_name=save_uploaded_file(uploaded,'idcard')
                if not file_name:
                    conn.close(); flash('ไฟล์บัตรประชาชนต้องเป็น JPG PNG WEBP หรือ PDF'); return redirect(request.url)
                if old_file: delete_uploaded_file(old_file)
            if name:
                if file_name: kept_files.add(file_name)
                updated_members.append((name,idcard,file_name))
        if not incomplete_import_allowed(event,team_name,updated_members):
            conn.close(); flash('กรุณากรอกชื่อนักกีฬาอย่างน้อย 1 คน หรือระบุชื่อทีมสำหรับรายการคู่/ทีม'); return redirect(request.url)
        for file_name in old_files.values():
            if file_name not in kept_files: delete_uploaded_file(file_name)
        is_complete=registration_is_complete(member_count,updated_members)
        status=request.form.get('status','pending')
        if status not in {'pending','approved'}: status='pending'
        if status=='approved' and not is_complete: status='pending'
        is_waitlist=1 if request.form.get('is_waitlist') else 0
        conn.execute('DELETE FROM certificates WHERE registration_id=?',(registration_id,))
        conn.execute('DELETE FROM registration_members WHERE registration_id=?',(registration_id,))
        conn.execute('''UPDATE registrations SET event_id=?,team_name=?,affiliation=?,contact_name=?,phone=?,notes=?,member_count=?,status=?,approved_at=?,is_waitlist=?,is_complete=?,coach_name=? WHERE id=?''',(event_id,team_name or None,affiliation or None,contact,phone,notes,member_count,status,now() if status=='approved' else None,is_waitlist,is_complete,coach_name or None,registration_id))
        for m in updated_members: conn.execute('INSERT INTO registration_members(registration_id,member_name,member_idcard,idcard_file) VALUES(?,?,?,?)',(registration_id,*m))
        conn.commit(); conn.close(); flash('แก้ไขข้อมูลผู้สมัครเรียบร้อยแล้ว' + ('' if is_complete else ' — รายชื่อนักกีฬายังไม่ครบ สามารถกลับมาแก้เพิ่มได้'))
        return redirect(url_for('tournament_registrations',tournament_id=reg['tournament_id'],event_id=event_id))
    max_slots=max([len(members), reg['member_count'] or 1] + [max(allowed_member_counts(e['category_type'], e['gender_type'], e['fixed_member_count'] if is_certificate_only_event(e) else None)) for e in events])
    member_slots=[]
    for i in range(max_slots): member_slots.append(members[i] if i<len(members) else None)
    conn.close()
    return render_template('edit_registration.html',reg=reg,events=events,member_slots=member_slots)

@app.route('/admin/registration/<int:registration_id>/award',methods=['POST'])
def update_registration_award(registration_id):
    if not is_logged_in(): return redirect(url_for('login'))
    conn=get_db(); r=conn.execute('''SELECT r.*,e.tournament_id,t.created_by FROM registrations r JOIN events e ON r.event_id=e.id JOIN tournaments t ON e.tournament_id=t.id WHERE r.id=?''',(registration_id,)).fetchone()
    if not r or r['created_by']!=session['user_id']: conn.close(); return redirect(url_for('admin_dashboard'))
    allowed={'participant','champion','runner_up_1','runner_up_2','runner_up_3','honorable','custom'}
    result=request.form.get('award_result','participant')
    if result not in allowed: result='participant'
    custom=request.form.get('award_custom','').strip() if result=='custom' else None
    if result=='custom' and not custom:
        conn.close(); flash('กรุณากรอกชื่อรางวัลพิเศษ'); return redirect(url_for('tournament_registrations',tournament_id=r['tournament_id'],event_id=r['event_id']))
    conn.execute('UPDATE registrations SET award_result=?,award_custom=?,award_updated_at=? WHERE id=?',(result,custom,now(),registration_id))
    conn.commit(); conn.close(); flash('บันทึกผลการแข่งขันเรียบร้อยแล้ว')
    return redirect(url_for('tournament_registrations',tournament_id=r['tournament_id'],event_id=r['event_id']))

@app.route('/admin/event/<int:event_id>/template')
def registration_template(event_id):
    if not is_logged_in(): return redirect(url_for('login'))
    conn=get_db(); e=get_owned_event(conn,event_id); conn.close()
    if not e or e['created_by']!=session['user_id']: abort(404)
    wb=openpyxl.Workbook(); ws=wb.active; ws.title='รายชื่อสมัคร'; headers=['ชื่อทีม','ต้นสังกัด','ผู้ติดต่อ','เบอร์โทร','จำนวนผู้เล่น','สมาชิก 1','เลขบัตร 1','สมาชิก 2','เลขบัตร 2','สมาชิก 3','เลขบัตร 3','สมาชิก 4','เลขบัตร 4','หมายเหตุ']; ws.append(headers); ws.append(['ตัวอย่างทีม A','โรงเรียน/ชมรม','นายผู้ติดต่อ','0812345678',event_suggested_member_count(e),'ชื่อสมาชิก 1','','ชื่อสมาชิก 2','','ชื่อสมาชิก 3','','ชื่อสมาชิก 4','',''])
    style_ws(ws); return excel_response(wb,f'event_{event_id}_registration_template.xlsx')

@app.route('/admin/event/<int:event_id>/import',methods=['GET','POST'])
def import_registrations(event_id):
    if not is_logged_in(): return redirect(url_for('login'))
    conn=get_db(); e=get_owned_event(conn,event_id)
    if not e or e['created_by']!=session['user_id']: conn.close(); abort(404)
    errors=[]; imported=0
    if request.method=='POST':
        f=request.files.get('excel_file')
        if not f or not allowed_file(f.filename,EXCEL_EXTENSIONS): conn.close(); flash('กรุณาเลือกไฟล์ .xlsx'); return redirect(request.url)
        wb=openpyxl.load_workbook(f, data_only=True); ws=wb.active
        for rowno, row in enumerate(ws.iter_rows(min_row=2,values_only=True), start=2):
            if not any(v not in (None,'') for v in row): continue
            vals=list(row)+['']*14; team,aff,contact,phone,count= [str(vals[i] or '').strip() for i in range(5)]
            try: count=int(float(count))
            except: errors.append(f'แถว {rowno}: จำนวนผู้เล่นไม่ถูกต้อง'); continue
            if count not in event_allowed_member_counts(e): errors.append(f'แถว {rowno}: จำนวนผู้เล่นไม่ตรงประเภท {category_label(e["category_type"])}'); continue
            members=[]
            for i in range(4):
                name=str(vals[5+i*2] or '').strip(); idc=str(vals[6+i*2] or '').strip()
                if name: members.append((name,idc))
            if not contact or not phone: errors.append(f'แถว {rowno}: กรุณากรอกผู้ติดต่อและเบอร์โทร'); continue
            if len(members)>count or not incomplete_import_allowed(e,team,members): errors.append(f'แถว {rowno}: รายชื่อนักกีฬาไม่ถูกต้อง'); continue
            is_complete=registration_is_complete(count,members)
            full, wait=registration_capacity_state(e)
            if full and not wait: errors.append(f'แถว {rowno}: จำนวนรับสมัครเต็มแล้ว'); continue
            code=unique_registration_code(conn); rid=insert_returning_id(conn,'''INSERT INTO registrations(event_id,team_name,affiliation,contact_name,phone,notes,created_at,member_count,source,registration_code,status,is_waitlist,is_complete) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)''',(event_id,team or None,aff or None,contact,phone,str(vals[13] or '').strip(),now(),count,'excel',code,'pending',1 if full and wait else 0,is_complete)); c=conn.cursor()
            for m in members: c.execute('INSERT INTO registration_members(registration_id,member_name,member_idcard) VALUES(?,?,?)',(rid,*m))
            imported+=1
        conn.execute('INSERT INTO import_logs(tournament_id,event_id,import_type,filename,imported_count,error_count,created_at) VALUES(?,?,?,?,?,?,?)',(e['tournament_id'],event_id,'registrations',secure_filename(f.filename),imported,len(errors),now())); conn.commit(); conn.close(); return render_template('import_result.html',title='ผลนำเข้ารายชื่อ',imported=imported,errors=errors,back_url=url_for('manage_events',tournament_id=e['tournament_id']))
    conn.close(); return render_template('import_excel.html',title='นำเข้าทีมหรือนักกีฬา',description=f'อีเวนต์: {event_display_name(e)}',template_url=url_for('registration_template',event_id=event_id))


def build_bulk_registration_workbook(events):
    """สร้าง Excel รายชื่อรวม ใช้ได้ทั้งฝั่งแอดมินและผู้สมัครทั่วไป"""
    wb=openpyxl.Workbook(); ws=wb.active; ws.title='รายชื่อสมัครรวม'
    headers=['รหัสอีเวนต์','ชื่ออีเวนต์','ชื่อทีม','ต้นสังกัด','ผู้ติดต่อ','เบอร์โทร','จำนวนผู้เล่น','สมาชิก 1','เลขบัตร 1','สมาชิก 2','เลขบัตร 2','สมาชิก 3','เลขบัตร 3','สมาชิก 4','เลขบัตร 4','หมายเหตุ']
    ws.append(headers)
    for e in events:
        ws.append([e['id'],event_display_name(e),'','','','',event_suggested_member_count(e),'','','','','','','','',''])
    for _ in range(max(50, len(events)*3)):
        ws.append(['','','','','','','','','','','','','','','',''])
    ref=wb.create_sheet('รายการอีเวนต์')
    ref.append(['รหัสอีเวนต์','ชื่ออีเวนต์','ประเภท','เพศ','รุ่น','จำนวนผู้เล่นที่รับได้'])
    for e in events:
        ref.append([e['id'],event_display_name(e),category_label(e['category_type']),gender_label(e['gender_type']),age_label(e['age_group']),','.join(str(n) for n in event_allowed_member_counts(e))])
    guide=wb.create_sheet('วิธีใช้')
    guide.append(['วิธีกรอกรายชื่อสมัครรวมทุกอีเวนต์'])
    guide.append(['1. กรอกข้อมูลในชีต “รายชื่อสมัครรวม” เพียงชีตเดียว'])
    guide.append(['2. ใช้รหัสอีเวนต์ตามชีต “รายการอีเวนต์” ระบบจะแยกผู้สมัครให้อัตโนมัติ'])
    guide.append(['3. หากมีหลายทีมในอีเวนต์เดียว ให้คัดลอกแถวของอีเวนต์นั้นเพิ่ม'])
    guide.append(['4. เดี่ยวรับ 1 คน, คู่ชาย/คู่หญิงรับ 2 หรือ 3 คน, คู่ผสมรับเฉพาะ 2 คน, ทีมรับ 3 หรือ 4 คน'])
    guide.append(['5. ชื่อทีมบังคับเฉพาะประเภททีม แต่กรอกได้ทุกประเภท'])
    guide.append(['6. ผู้ติดต่อและเบอร์โทรกรอกใน Excel หรือกรอกค่าเริ่มต้นตอนอัปโหลดก็ได้'])
    guide.append(['7. รายการคู่และทีมสามารถเว้นชื่อนักกีฬาบางคนไว้ก่อน แล้วให้ผู้ดูแลเติมภายหลังได้'])
    guide.append(['8. ห้ามแก้ชื่อหัวตาราง และใช้ไฟล์ .xlsx เท่านั้น'])
    style_ws(ws); style_ws(ref)
    guide.column_dimensions['A'].width=110
    ws.freeze_panes='A2'; ws.auto_filter.ref=ws.dimensions
    return wb

@app.route('/admin/tournament/<int:tournament_id>/registration-template-all')
def bulk_registration_template(tournament_id):
    """ดาวน์โหลด Excel ไฟล์เดียวสำหรับกรอกรายชื่อผู้สมัครทุกอีเวนต์ในงาน"""
    if not is_logged_in(): return redirect(url_for('login'))
    conn=get_db(); t=get_owned_tournament(conn,tournament_id)
    if not t: conn.close(); abort(404)
    events=conn.execute('SELECT * FROM events WHERE tournament_id=? ORDER BY age_group,category_type,gender_type,id',(tournament_id,)).fetchall(); conn.close()
    if not events:
        flash('กรุณาสร้างอีเวนต์ก่อนดาวน์โหลดไฟล์ตัวอย่างรายชื่อรวม')
        return redirect(url_for('manage_events',tournament_id=tournament_id))
    return excel_response(build_bulk_registration_workbook(events),f'tournament_{tournament_id}_all_registration_template.xlsx')

def _event_lookup_for_bulk_import(events):
    by_id={int(e['id']):e for e in events}
    by_name={}
    for e in events:
        name=event_display_name(e).strip()
        by_name.setdefault(name,[]).append(e)
    return by_id,by_name


def _int_from_excel(value, default=None):
    if value in (None,''): return default
    try: return int(float(value))
    except (TypeError,ValueError): return default


def delete_tournament_registrations(conn, tournament_id):
    """ลบรายชื่อสมัครเดิมทั้งหมดของงานแข่งขันนี้ ก่อนนำเข้าไฟล์ใหม่แบบแทนข้อมูลเดิม"""
    rows = conn.execute('''
        SELECT r.id
        FROM registrations r
        JOIN events e ON r.event_id = e.id
        WHERE e.tournament_id = ?
    ''', (tournament_id,)).fetchall()
    ids = [row['id'] for row in rows]
    if not ids:
        return 0
    marks = ','.join(['?'] * len(ids))
    params = tuple(ids)
    conn.execute(f'DELETE FROM certificates WHERE registration_id IN ({marks})', params)
    conn.execute(f'DELETE FROM registration_members WHERE registration_id IN ({marks})', params)
    conn.execute(f'DELETE FROM registrations WHERE id IN ({marks})', params)
    return len(ids)


@app.route('/admin/tournament/<int:tournament_id>/registration-import-all',methods=['GET','POST'])
def import_registrations_all(tournament_id):
    """Import รายชื่อจาก Excel ไฟล์เดียว แล้วแยกผู้สมัครเข้าทุกอีเวนต์อัตโนมัติ

    โหมดใหม่:
    - เพิ่มต่อจากเดิม: ใช้เหมือนเดิม
    - แทนข้อมูลเดิม: ตรวจไฟล์ก่อน ถ้าไม่มีปัญหาจึงลบรายชื่อเดิมของงานนี้และนำเข้าใหม่ทั้งหมด
    """
    if not is_logged_in(): return redirect(url_for('login'))
    conn=get_db(); t=get_owned_tournament(conn,tournament_id)
    if not t: conn.close(); abort(404)
    events=conn.execute('SELECT * FROM events WHERE tournament_id=? ORDER BY id',(tournament_id,)).fetchall()
    if not events:
        conn.close(); flash('กรุณาสร้างอีเวนต์ก่อนนำเข้ารายชื่อรวม'); return redirect(url_for('manage_events',tournament_id=tournament_id))
    errors=[]; imported=0; imported_by_event={}; deleted_count=0
    if request.method=='POST':
        replace_existing = request.form.get('replace_existing') == '1'
        f=request.files.get('excel_file')
        if not f or not allowed_file(f.filename,EXCEL_EXTENSIONS):
            conn.close(); flash('กรุณาเลือกไฟล์ .xlsx'); return redirect(request.url)
        try:
            wb=openpyxl.load_workbook(f,data_only=True)
            ws=wb['รายชื่อสมัครรวม'] if 'รายชื่อสมัครรวม' in wb.sheetnames else wb.active
        except Exception:
            conn.close(); flash('ไม่สามารถอ่านไฟล์ Excel ได้ กรุณาใช้ไฟล์ .xlsx ที่ดาวน์โหลดจากระบบ'); return redirect(request.url)

        by_id,by_name=_event_lookup_for_bulk_import(events)
        event_map={int(e['id']):e for e in events}
        records=[]
        for rowno,row in enumerate(ws.iter_rows(min_row=2,values_only=True),start=2):
            if not any(v not in (None,'') for v in row): continue
            vals=list(row)+['']*16
            event_id=_int_from_excel(vals[0])
            event_name=str(vals[1] or '').strip()
            e=by_id.get(event_id) if event_id is not None else None
            if not e and event_name:
                matches=by_name.get(event_name,[])
                if len(matches)==1: e=matches[0]
                elif len(matches)>1:
                    errors.append(f'แถว {rowno}: ชื่ออีเวนต์ซ้ำกัน กรุณาระบุรหัสอีเวนต์'); continue
            if not e:
                errors.append(f'แถว {rowno}: ไม่พบอีเวนต์ กรุณาตรวจรหัสหรือชื่ออีเวนต์'); continue
            team=str(vals[2] or '').strip(); aff=str(vals[3] or '').strip(); contact=str(vals[4] or '').strip(); phone=str(vals[5] or '').strip(); count=_int_from_excel(vals[6])
            if count is None:
                errors.append(f'แถว {rowno}: จำนวนผู้เล่นไม่ถูกต้อง'); continue
            if count not in event_allowed_member_counts(e):
                allowed='/'.join(str(n) for n in event_allowed_member_counts(e))
                errors.append(f'แถว {rowno}: {event_display_name(e)} รับผู้เล่น {allowed} คน'); continue
            if e['category_type']=='team' and not team:
                errors.append(f'แถว {rowno}: ประเภททีมต้องกรอกชื่อทีม'); continue
            members=[]
            for i in range(4):
                name=str(vals[7+i*2] or '').strip(); idc=str(vals[8+i*2] or '').strip()
                if name: members.append((name,idc))
            if not contact or not phone:
                errors.append(f'แถว {rowno}: กรุณากรอกผู้ติดต่อและเบอร์โทร'); continue
            if len(members)>count or not incomplete_import_allowed(e,team,members):
                errors.append(f'แถว {rowno}: รายชื่อนักกีฬาไม่ถูกต้อง'); continue
            records.append({'rowno':rowno,'event_id':int(e['id']),'event':e,'event_name':event_display_name(e),'team_name':team,'affiliation':aff,'contact_name':contact,'phone':phone,'member_count':count,'members':members,'is_complete':registration_is_complete(count,members),'notes':str(vals[15] or '').strip()})

        # ตรวจจำนวนรับสมัครแบบรวมทั้งไฟล์ก่อนบันทึกจริง
        planned_active={}; planned_wait={}; accepted_records=[]
        current_counts=event_counts_map(conn,event_ids=[e['id'] for e in events]) if not replace_existing else {}
        for record in records:
            e=event_map.get(int(record['event_id']))
            if not e:
                errors.append(f"แถว {record['rowno']}: ไม่พบอีเวนต์"); continue
            eid=int(e['id'])
            if eid not in planned_active:
                if replace_existing:
                    planned_active[eid]=0; planned_wait[eid]=0
                else:
                    st=current_counts.get(eid,_zero_event_count())
                    planned_active[eid]=st['active']
                    planned_wait[eid]=st['waitlist']
            has_limit=bool(e['has_limit'] and int(e['max_slots'] or 0)>0)
            if not has_limit or planned_active[eid]<int(e['max_slots'] or 0):
                record['is_waitlist']=0; planned_active[eid]+=1; accepted_records.append(record); continue
            can_wait=bool(e['waitlist_enabled'] and (int(e['waitlist_limit'] or 0)<=0 or planned_wait[eid]<int(e['waitlist_limit'] or 0)))
            if can_wait:
                record['is_waitlist']=1; planned_wait[eid]+=1; accepted_records.append(record); continue
            errors.append(f"แถว {record['rowno']}: {record['event_name']} เต็มแล้ว")

        if replace_existing and errors:
            conn.close()
            errors.insert(0,'ยังไม่ได้ลบข้อมูลเดิม เพราะไฟล์ที่อัปโหลดยังพบปัญหา กรุณาแก้ไฟล์แล้วอัปโหลดใหม่')
            return render_template('import_result.html',title='ผลตรวจไฟล์นำเข้ารายชื่อรวมทุกอีเวนต์',imported=0,errors=errors,event_summary=[],back_url=url_for('manage_events',tournament_id=tournament_id))

        if replace_existing:
            deleted_count=delete_tournament_registrations(conn,tournament_id)

        c=conn.cursor()
        for record in accepted_records:
            e=record['event']
            code=unique_registration_code(conn)
            rid=insert_returning_id(conn,'''INSERT INTO registrations(event_id,team_name,affiliation,contact_name,phone,notes,created_at,member_count,source,registration_code,status,is_waitlist,is_complete) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)''',(e['id'],record['team_name'] or None,record['affiliation'] or None,record['contact_name'],record['phone'],record['notes'],now(),record['member_count'],'excel_all_replace' if replace_existing else 'excel_all',code,'pending',record.get('is_waitlist',0),record['is_complete']))
            for m in record['members']:
                c.execute('INSERT INTO registration_members(registration_id,member_name,member_idcard) VALUES(?,?,?)',(rid,*m))
            imported+=1; imported_by_event[e['id']]=imported_by_event.get(e['id'],0)+1

        import_type='registrations_all_events_replace' if replace_existing else 'registrations_all_events'
        conn.execute('INSERT INTO import_logs(tournament_id,import_type,filename,imported_count,error_count,created_at) VALUES(?,?,?,?,?,?)',(tournament_id,import_type,secure_filename(f.filename),imported,len(errors),now()))
        conn.commit(); conn.close()
        event_summary=[]
        if replace_existing:
            event_summary.append('โหมดนำเข้า: แทนข้อมูลเดิมในงานนี้')
            event_summary.append(f'ลบข้อมูลเดิมก่อนนำเข้า: {deleted_count} รายการ')
        else:
            event_summary.append('โหมดนำเข้า: เพิ่มต่อจากข้อมูลเดิม')
        event_summary += [f"{event_display_name(e)}: {imported_by_event[e['id']]} รายการ" for e in events if imported_by_event.get(e['id'])]
        return render_template('import_result.html',title='ผลนำเข้ารายชื่อรวมทุกอีเวนต์',imported=imported,errors=errors,event_summary=event_summary,back_url=url_for('manage_events',tournament_id=tournament_id))
    conn.close()
    return render_template('import_excel.html',title='นำเข้ารายชื่อรวมทุกอีเวนต์',description=f'งานแข่งขัน: {t["title"]} — ใช้ Excel ไฟล์เดียว ระบบจะแยกรายชื่อเข้าทุกอีเวนต์ให้อัตโนมัติ',template_url=url_for('bulk_registration_template',tournament_id=tournament_id),allow_replace=True)

def _public_tournament_and_events(tournament_id):
    conn=get_db()
    tournament=conn.execute('SELECT * FROM tournaments WHERE id=?',(tournament_id,)).fetchone()
    if not tournament:
        conn.close(); return None,[]
    events=conn.execute('SELECT * FROM events WHERE tournament_id=? AND is_open=1 ORDER BY age_group,category_type,gender_type,id',(tournament_id,)).fetchall()
    conn.close(); return tournament,events


def _save_public_bulk_cache(payload):
    token=secrets.token_urlsafe(24)
    path=os.path.join(BULK_IMPORT_CACHE_FOLDER,token+'.json')
    with open(path,'w',encoding='utf-8') as f: json.dump(payload,f,ensure_ascii=False)
    return token


def _load_public_bulk_cache(token,delete=False):
    if not token or not all(ch.isalnum() or ch in '-_' for ch in token): return None
    path=os.path.join(BULK_IMPORT_CACHE_FOLDER,token+'.json')
    if not os.path.exists(path): return None
    try:
        if time.time()-os.path.getmtime(path)>6*60*60:
            os.remove(path); return None
        with open(path,'r',encoding='utf-8') as f: payload=json.load(f)
        if delete: os.remove(path)
        return payload
    except Exception:
        return None


def _parse_public_bulk_sheet(events,ws,default_contact='',default_phone=''):
    by_id,by_name=_event_lookup_for_bulk_import(events)
    errors=[]; records=[]
    for rowno,row in enumerate(ws.iter_rows(min_row=2,values_only=True),start=2):
        if not any(v not in (None,'') for v in row): continue
        vals=list(row)+['']*16
        event_id=_int_from_excel(vals[0]); event_name=str(vals[1] or '').strip()
        e=by_id.get(event_id) if event_id is not None else None
        if not e and event_name:
            matches=by_name.get(event_name,[])
            if len(matches)==1: e=matches[0]
            elif len(matches)>1:
                errors.append(f'แถว {rowno}: ชื่ออีเวนต์ซ้ำกัน กรุณาระบุรหัสอีเวนต์'); continue
        if not e:
            errors.append(f'แถว {rowno}: ไม่พบอีเวนต์ที่เปิดรับสมัคร กรุณาตรวจรหัสหรือชื่ออีเวนต์'); continue
        team=str(vals[2] or '').strip(); aff=str(vals[3] or '').strip()
        contact=str(vals[4] or '').strip() or default_contact
        phone=str(vals[5] or '').strip() or default_phone
        count=_int_from_excel(vals[6])
        if count is None:
            errors.append(f'แถว {rowno}: จำนวนผู้เล่นไม่ถูกต้อง'); continue
        if count not in event_allowed_member_counts(e):
            allowed='/'.join(str(n) for n in event_allowed_member_counts(e))
            errors.append(f'แถว {rowno}: {event_display_name(e)} รับผู้เล่น {allowed} คน'); continue
        if e['category_type']=='team' and not team:
            errors.append(f'แถว {rowno}: ประเภททีมต้องกรอกชื่อทีม'); continue
        members=[]
        for i in range(4):
            name=str(vals[7+i*2] or '').strip(); idc=str(vals[8+i*2] or '').strip()
            if name: members.append({'name':name,'idcard':idc})
        if len(members)>count or not incomplete_import_allowed(e,team,members):
            errors.append(f'แถว {rowno}: รายชื่อนักกีฬาไม่ถูกต้อง'); continue
        if not contact or not phone:
            errors.append(f'แถว {rowno}: กรุณากรอกผู้ติดต่อและเบอร์โทรใน Excel หรือกรอกค่าเริ่มต้นก่อนอัปโหลด'); continue
        records.append({'rowno':rowno,'event_id':int(e['id']),'event_name':event_display_name(e),'category_type':e['category_type'],'gender_type':e['gender_type'],'team_name':team,'affiliation':aff,'contact_name':contact,'phone':phone,'member_count':count,'members':members,'is_complete':registration_is_complete(count,members),'notes':str(vals[15] or '').strip()})
    return records,errors


def _preview_capacity(records,events):
    event_map={int(e['id']):e for e in events}; planned_active={}; planned_wait={}; accepted=[]; errors=[]
    conn=get_db(); current_counts=event_counts_map(conn,event_ids=[e['id'] for e in events]); conn.close()
    for record in records:
        e=event_map.get(int(record['event_id']))
        if not e or not e['is_open']:
            errors.append(f"แถว {record['rowno']}: อีเวนต์ปิดรับสมัครแล้ว"); continue
        eid=int(e['id'])
        if eid not in planned_active:
            st=current_counts.get(eid,_zero_event_count())
            planned_active[eid]=st['active']; planned_wait[eid]=st['waitlist']
        has_limit=bool(e['has_limit'] and int(e['max_slots'] or 0)>0)
        if not has_limit or planned_active[eid]<int(e['max_slots'] or 0):
            record['is_waitlist']=0; planned_active[eid]+=1; accepted.append(record); continue
        can_wait=bool(e['waitlist_enabled'] and (int(e['waitlist_limit'] or 0)<=0 or planned_wait[eid]<int(e['waitlist_limit'] or 0)))
        if can_wait:
            record['is_waitlist']=1; planned_wait[eid]+=1; accepted.append(record)
        else:
            errors.append(f"แถว {record['rowno']}: {record['event_name']} เต็มแล้ว")
    return accepted,errors


@app.route('/tournament/<int:tournament_id>/bulk-registration-template')
def public_bulk_registration_template(tournament_id):
    tournament,events=_public_tournament_and_events(tournament_id)
    if not tournament: abort(404)
    if not tournament['is_open'] or not events:
        flash('งานนี้ยังไม่มีอีเวนต์ที่เปิดรับสมัคร'); return redirect(url_for('home'))
    return excel_response(build_bulk_registration_workbook(events),f'tournament_{tournament_id}_public_bulk_registration_template.xlsx')


@app.route('/tournament/<int:tournament_id>/bulk-register',methods=['GET','POST'])
def public_bulk_register(tournament_id):
    tournament,events=_public_tournament_and_events(tournament_id)
    if not tournament: abort(404)
    if not tournament['is_open']:
        flash('งานแข่งขันนี้ปิดรับสมัครแล้ว'); return redirect(url_for('home'))
    if not events:
        flash('ยังไม่มีอีเวนต์ที่เปิดรับสมัคร'); return redirect(url_for('home'))
    if request.method=='POST':
        action=request.form.get('action','preview')
        if action=='confirm':
            token=request.form.get('token','')
            payload=_load_public_bulk_cache(token,delete=True)
            if not payload or int(payload.get('tournament_id',0))!=tournament_id:
                flash('ข้อมูลพรีวิวหมดอายุ กรุณาอัปโหลด Excel ใหม่'); return redirect(request.url)
            records=payload.get('records',[]); conn=get_db(); imported=[]; errors=[]; by_event={}
            event_map={int(e['id']):e for e in events}
            current_counts=event_counts_map(conn,event_ids=[e['id'] for e in events])
            planned_active={}; planned_wait={}; c=conn.cursor()
            for record in records:
                e=event_map.get(int(record['event_id']))
                if not e:
                    errors.append(f"แถว {record['rowno']}: อีเวนต์ปิดรับสมัครแล้ว"); continue
                eid=int(e['id'])
                if eid not in planned_active:
                    st=current_counts.get(eid,_zero_event_count())
                    planned_active[eid]=st['active']; planned_wait[eid]=st['waitlist']
                has_limit=bool(e['has_limit'] and int(e['max_slots'] or 0)>0)
                is_waitlist=0
                if has_limit and planned_active[eid]>=int(e['max_slots'] or 0):
                    can_wait=bool(e['waitlist_enabled'] and (int(e['waitlist_limit'] or 0)<=0 or planned_wait[eid]<int(e['waitlist_limit'] or 0)))
                    if not can_wait:
                        errors.append(f"แถว {record['rowno']}: {record['event_name']} เต็มแล้ว"); continue
                    is_waitlist=1; planned_wait[eid]+=1
                else:
                    planned_active[eid]+=1
                code=unique_registration_code(conn)
                rid=insert_returning_id(conn,'''INSERT INTO registrations(event_id,team_name,affiliation,contact_name,phone,notes,created_at,member_count,source,registration_code,status,is_waitlist,is_complete) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)''',(e['id'],record['team_name'] or None,record['affiliation'] or None,record['contact_name'],record['phone'],record['notes'],now(),record['member_count'],'public_excel_all',code,'pending',is_waitlist,record.get('is_complete',0)))
                for m in record['members']: c.execute('INSERT INTO registration_members(registration_id,member_name,member_idcard) VALUES(?,?,?)',(rid,m['name'],m['idcard']))
                item=dict(record); item.update(registration_code=code,registration_id=rid,is_waitlist=is_waitlist)
                imported.append(item); by_event[item['event_name']]=by_event.get(item['event_name'],0)+1
            conn.execute('INSERT INTO import_logs(tournament_id,import_type,filename,imported_count,error_count,created_at) VALUES(?,?,?,?,?,?)',(tournament_id,'public_registrations_all_events',payload.get('filename','public_bulk.xlsx'),len(imported),len(errors),now()))
            conn.commit(); conn.close()
            return render_template('public_bulk_import_result.html',tournament=tournament,imported=imported,errors=errors,event_summary=sorted(by_event.items()))
        f=request.files.get('excel_file')
        if not f or not allowed_file(f.filename,EXCEL_EXTENSIONS):
            flash('กรุณาเลือกไฟล์ .xlsx'); return redirect(request.url)
        try:
            wb=openpyxl.load_workbook(f,data_only=True)
            ws=wb['รายชื่อสมัครรวม'] if 'รายชื่อสมัครรวม' in wb.sheetnames else wb.active
        except Exception:
            flash('ไม่สามารถอ่านไฟล์ Excel ได้ กรุณาใช้ไฟล์ตัวอย่างที่ดาวน์โหลดจากระบบ'); return redirect(request.url)
        records,errors=_parse_public_bulk_sheet(events,ws,request.form.get('default_contact','').strip(),request.form.get('default_phone','').strip())
        records,capacity_errors=_preview_capacity(records,events); errors.extend(capacity_errors)
        token=_save_public_bulk_cache({'tournament_id':tournament_id,'filename':secure_filename(f.filename),'records':records}) if records else None
        return render_template('public_bulk_import.html',tournament=tournament,events=events,records=records,errors=errors,token=token,default_contact=request.form.get('default_contact','').strip(),default_phone=request.form.get('default_phone','').strip())
    return render_template('public_bulk_import.html',tournament=tournament,events=events,records=None,errors=[],token=None)

@app.route('/admin/tournament/<int:tournament_id>/event-template')
def event_template(tournament_id):
    if not is_logged_in(): return redirect(url_for('login'))
    conn=get_db(); t=get_owned_tournament(conn,tournament_id); conn.close()
    if not t: abort(404)
    wb=openpyxl.Workbook(); ws=wb.active; ws.title='อีเวนต์'
    ws.append(['รูปแบบอีเวนต์','ชื่ออีเวนต์','ชนิดกีฬา','ประเภท','เพศ','รุ่น','จำนวนนักกีฬาในเกียรติบัตร','เปิดรับสมัคร','มีค่าสมัคร','ค่าสมัคร','คิดราคาต่อ','ต้องแนบสลิป','จำกัดจำนวน','จำนวนสูงสุด','เปิดสำรอง','จำนวนสำรองสูงสุด'])
    ws.append(['ออกเกียรติบัตรอย่างเดียว','ฟุตบอลชาย รุ่นอายุไม่เกิน 14 ปี','ฟุตบอล','ทีม','ชาย','รุ่นอายุไม่เกิน 14 ปี',11,'ใช่','ไม่',0,'ทีม','ไม่','ไม่',0,'ไม่',0])
    ws.append(['สมัครแข่งขันปกติ','คู่ชายทั่วไป','เปตอง','คู่','ชาย','ทั่วไป','','ใช่','ใช่',100,'ทีม','ไม่','ใช่',24,'ใช่',5])
    guide=wb.create_sheet('วิธีใช้')
    guide.append(['รูปแบบอีเวนต์: ใส่ “ออกเกียรติบัตรอย่างเดียว” เมื่อต้องการให้กรอกชื่อแล้วปริ้นเกียรติบัตรได้ทันที หรือ “สมัครแข่งขันปกติ” สำหรับรับสมัครแข่งขันแบบเดิม'])
    guide.append(['โหมดเกียรติบัตรอย่างเดียว ต้องระบุชนิดกีฬา ประเภท เพศ รุ่น และจำนวนนักกีฬาในเกียรติบัตร'])
    style_ws(ws); guide.column_dimensions['A'].width=120
    return excel_response(wb,f'tournament_{tournament_id}_event_template.xlsx')

def normalize_event_mode(v):
    raw=str(v or '').strip().lower()
    if raw in {'ออกเกียรติบัตรอย่างเดียว','เกียรติบัตร','certificate_only','certificate only','cert'}:
        return 'certificate_only'
    return 'competition'

@app.route('/admin/tournament/<int:tournament_id>/event-import',methods=['GET','POST'])
def import_events(tournament_id):
    if not is_logged_in(): return redirect(url_for('login'))
    conn=get_db(); t=get_owned_tournament(conn,tournament_id)
    if not t: conn.close(); abort(404)
    errors=[]; imported=0
    if request.method=='POST':
        f=request.files.get('excel_file')
        if not f or not allowed_file(f.filename,EXCEL_EXTENSIONS): conn.close(); flash('กรุณาเลือกไฟล์ .xlsx'); return redirect(request.url)
        ws=openpyxl.load_workbook(f,data_only=True).active
        header=[str(c.value or '').strip() for c in ws[1]]
        new_format=bool(header and header[0]=='รูปแบบอีเวนต์')
        for rowno,row in enumerate(ws.iter_rows(min_row=2,values_only=True),start=2):
            if not any(v not in (None,'') for v in row): continue
            if new_format:
                v=list(row)+['']*16
                event_mode=normalize_event_mode(v[0]); event_name=str(v[1] or '').strip(); sport_name=str(v[2] or '').strip(); cat=normalize_category(v[3]); gender=normalize_gender(v[4]); age=normalize_age(v[5]); fixed_count=parse_int(v[6] or 0,0,0,99) if event_mode=='certificate_only' else 0
                open_value=v[7]; hasfee=truthy(v[8]); fee_value=v[9]; fee_per_value=v[10]; slip_value=v[11]; haslimit=truthy(v[12]); limit_value=v[13]; wait_value=v[14]; waitlimit_value=v[15]
            else:
                v=list(row)+['']*13
                event_mode='competition'; event_name=str(v[0] or '').strip(); sport_name=''; cat=normalize_category(v[1]); gender=normalize_gender(v[2]); age=normalize_age(v[3]); fixed_count=0
                open_value=v[4]; hasfee=truthy(v[5]); fee_value=v[6]; fee_per_value=v[7]; slip_value=v[8]; haslimit=truthy(v[9]); limit_value=v[10]; wait_value=v[11]; waitlimit_value=v[12]
            if not cat: errors.append(f'แถว {rowno}: ประเภทต้องเป็น เดี่ยว คู่ หรือ ทีม'); continue
            try:
                fee=parse_int(fee_value or 0,0,0); limit=parse_int(limit_value or 0,0,0); waitlimit=parse_int(waitlimit_value or 0,0,0)
            except Exception:
                errors.append(f'แถว {rowno}: จำนวนเงินหรือจำนวนรับสมัครไม่ถูกต้อง'); continue
            team_size=fixed_count if fixed_count > 0 else suggested_member_count(cat,gender)
            conn.execute('''INSERT INTO events(tournament_id,event_name,category_type,gender_type,age_group,max_slots,fee,team_size,is_open,created_at,has_fee,fee_per,require_slip,has_limit,waitlist_enabled,waitlist_limit,event_mode,sport_name,fixed_member_count) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',(tournament_id,event_name,cat,gender,age,limit if haslimit else 0,fee if hasfee else 0,team_size,1 if truthy(open_value) else 0,now(),1 if hasfee else 0,'person' if str(fee_per_value or '').strip() in {'คน','person'} else 'team',1 if hasfee and truthy(slip_value) else 0,1 if haslimit else 0,1 if haslimit and truthy(wait_value) else 0,waitlimit,event_mode,sport_name,fixed_count)); imported+=1
        conn.execute('INSERT INTO import_logs(tournament_id,import_type,filename,imported_count,error_count,created_at) VALUES(?,?,?,?,?,?)',(tournament_id,'events',secure_filename(f.filename),imported,len(errors),now())); conn.commit(); conn.close(); return render_template('import_result.html',title='ผลนำเข้าอีเวนต์',imported=imported,errors=errors,back_url=url_for('manage_events',tournament_id=tournament_id))
    conn.close(); return render_template('import_excel.html',title='นำเข้าอีเวนต์จาก Excel',description=f'งานแข่งขัน: {t["title"]}',template_url=url_for('event_template',tournament_id=tournament_id))


@app.route('/admin/tournament/<int:tournament_id>/certificate-settings',methods=['GET','POST'])
def certificate_settings(tournament_id):
    if not is_logged_in(): return redirect(url_for('login'))
    conn=get_db(); t=get_owned_tournament(conn,tournament_id)
    if not t: conn.close(); abort(404)
    if request.method=='POST':
        values={
            'certificates_enabled':1 if request.form.get('certificates_enabled') else 0,
            'certificate_self_download':1 if request.form.get('certificate_self_download') else 0,
            'certificate_require_approval':1 if request.form.get('certificate_require_approval') else 0,
            'cert_org':request.form.get('cert_org','').strip(),
            'cert_date':request.form.get('cert_date','').strip(),
            'cert_place':request.form.get('cert_place','').strip(),
            'cert_signer':request.form.get('cert_signer','').strip(),
            'cert_signer_position':request.form.get('cert_signer_position','').strip(),
            'cert_style':request.form.get('cert_style','navy_gold') if request.form.get('cert_style') in {'navy_gold','classic_gold'} else 'navy_gold',
            'cert_heading':request.form.get('cert_heading','').strip(),
            'cert_footer_note':request.form.get('cert_footer_note','').strip(),
        }
        asset_fields=['cert_logo_1','cert_logo_2','cert_logo_3','cert_background','cert_signature','cert_stamp']
        for field in asset_fields:
            current=t[field]
            if request.form.get(f'remove_{field}'):
                delete_uploaded_file(current); current=None
            f=request.files.get(field)
            if f and f.filename:
                if field == 'cert_background' and not allowed_file(f.filename, {'png'}):
                    conn.close(); flash('เทมเพลตพื้นหลังเกียรติบัตรต้องเป็นไฟล์ PNG เท่านั้น'); return redirect(request.url)
                saved=save_certificate_asset(f,field)
                if not saved:
                    conn.close(); flash('ไฟล์โลโก้ ลายเซ็น และตราประทับ ต้องเป็น PNG JPG JPEG หรือ WEBP'); return redirect(request.url)
                delete_uploaded_file(current); current=saved
            values[field]=current
        sets=','.join([f'{key}=?' for key in values])
        conn.execute(f'UPDATE tournaments SET {sets} WHERE id=?',(*values.values(),tournament_id))
        conn.commit(); conn.close(); flash('บันทึกตั้งค่าและรูปแบบเกียรติบัตรแล้ว'); return redirect(url_for('certificate_settings',tournament_id=tournament_id))
    conn.close(); return render_template('certificate_settings.html',tournament=t)

@app.route('/admin/tournament/<int:tournament_id>/certificate-preview')
def certificate_preview(tournament_id):
    if not is_logged_in(): return redirect(url_for('login'))
    conn=get_db()
    t=get_owned_tournament(conn,tournament_id)
    if not t:
        conn.close(); abort(404)

    # ใช้อีเวนต์จริงของงานนี้ในการพรีวิว เพื่อให้ชนิดกีฬา/รุ่น/ประเภทตรงกับที่แอดมินตั้งค่า
    event_id=request.args.get('event_id',type=int)
    if event_id:
        sample_event=conn.execute('SELECT * FROM events WHERE id=? AND tournament_id=?',(event_id,tournament_id)).fetchone()
    else:
        sample_event=conn.execute('SELECT * FROM events WHERE tournament_id=? ORDER BY id LIMIT 1',(tournament_id,)).fetchone()
    conn.close()

    if sample_event:
        reg=dict(t)
        reg.update(dict(sample_event))
        reg.update({
            'tournament_title':t['title'],
            'affiliation':'โรงเรียนตัวอย่าง',
            'coach_name':'นายผู้ฝึกสอน ตัวอย่าง',
            'award_result':'champion',
            'award_custom':'',
            'team_name':'ทีมตัวอย่าง',
        })
    else:
        reg=dict(t)
        reg.update({
            'tournament_title':t['title'], 'event_name':'ฟุตบอลชาย รุ่นอายุไม่เกิน 14 ปี',
            'category_type':'team','gender_type':'male','age_group':'รุ่นอายุไม่เกิน 14 ปี',
            'event_mode':'certificate_only','sport_name':'ฟุตบอล','fixed_member_count':11,
            'affiliation':'โรงเรียนตัวอย่าง','coach_name':'นายผู้ฝึกสอน ตัวอย่าง','award_result':'champion','award_custom':'',
            'team_name':'ทีมตัวอย่าง',
        })
    return render_template('certificate.html',reg=reg,member={'member_name':'นายตัวอย่าง ผู้เข้าแข่งขัน'},verification_code=None,team_mode=False,coach_mode=False,preview_mode=True)

def cert_access(reg):
    instant=is_certificate_only_event(reg)
    return bool(reg['certificates_enabled'] and (is_logged_in() or (reg['certificate_self_download'] and (instant or not reg['certificate_require_approval'] or reg['status']=='approved'))))

def get_or_create_cert(conn,registration_id,member_id=None,ctype='individual'):
    if member_id is None:
        r=conn.execute('SELECT * FROM certificates WHERE registration_id=? AND member_id IS NULL AND certificate_type=?',(registration_id,ctype)).fetchone()
    else:
        r=conn.execute('SELECT * FROM certificates WHERE registration_id=? AND member_id=? AND certificate_type=?',(registration_id,member_id,ctype)).fetchone()
    if r: return r['verification_code']
    code='CERT-'+datetime.now().strftime('%y')+'-'+secrets.token_hex(4).upper(); conn.execute('INSERT INTO certificates(registration_id,member_id,certificate_type,verification_code,issued_at) VALUES(?,?,?,?,?)',(registration_id,member_id,ctype,code,now())); conn.commit(); return code

def get_cert_data(registration_id,member_id=None,ctype='individual'):
    conn=get_db(); reg=conn.execute('''SELECT r.*,e.event_name,e.category_type,e.gender_type,e.age_group,e.event_mode,e.sport_name,e.fixed_member_count,t.title tournament_title,t.certificates_enabled,t.certificate_self_download,t.certificate_require_approval,t.cert_org,t.cert_date,t.cert_place,t.cert_signer,t.cert_signer_position,t.cert_style,t.cert_heading,t.cert_footer_note,t.cert_logo_1,t.cert_logo_2,t.cert_logo_3,t.cert_background,t.cert_signature,t.cert_stamp FROM registrations r JOIN events e ON r.event_id=e.id JOIN tournaments t ON e.tournament_id=t.id WHERE r.id=?''',(registration_id,)).fetchone()
    if not reg or not cert_access(reg) or (not is_certificate_only_event(reg) and not reg['is_complete']): conn.close(); abort(403)
    member=conn.execute('SELECT * FROM registration_members WHERE id=? AND registration_id=?',(member_id,registration_id)).fetchone() if member_id else None
    if member_id is not None and not member: conn.close(); abort(404)
    code=get_or_create_cert(conn,registration_id,member_id,ctype); conn.close(); return reg,member,code

@app.route('/certificate/<int:registration_id>/member/<int:member_id>')
def certificate_member(registration_id,member_id):
    reg,member,code=get_cert_data(registration_id,member_id,'individual')
    return render_template('certificate.html',reg=reg,member=member,verification_code=code,team_mode=False,coach_mode=False)

@app.route('/certificate/<int:registration_id>/team')
def certificate_team(registration_id):
    reg,member,code=get_cert_data(registration_id,None,'team')
    return render_template('certificate.html',reg=reg,member=None,verification_code=code,team_mode=True,coach_mode=False)

@app.route('/certificate/<int:registration_id>/coach')
def certificate_coach(registration_id):
    reg,member,code=get_cert_data(registration_id,None,'coach')
    if not (reg['coach_name'] or '').strip():
        abort(404)
    return render_template('certificate.html',reg=reg,member={'member_name':reg['coach_name']},verification_code=code,team_mode=False,coach_mode=True)

@app.route('/admin/event/<int:event_id>/certificates/print')
def print_event_certificates(event_id):
    if not is_logged_in(): return redirect(url_for('login'))
    conn=get_db(); event=get_owned_event(conn,event_id)
    if not event or event['created_by']!=session['user_id']: conn.close(); abort(404)
    if not event['certificates_enabled']:
        conn.close(); flash('กรุณาเปิดใช้งานเกียรติบัตรก่อน'); return redirect(url_for('certificate_settings',tournament_id=event['tournament_id']))
    regs=conn.execute('''SELECT r.*,e.event_name,e.category_type,e.gender_type,e.age_group,e.event_mode,e.sport_name,e.fixed_member_count,
        t.title tournament_title,t.certificates_enabled,t.certificate_self_download,t.certificate_require_approval,
        t.cert_org,t.cert_date,t.cert_place,t.cert_signer,t.cert_signer_position,t.cert_style,t.cert_heading,t.cert_footer_note,
        t.cert_logo_1,t.cert_logo_2,t.cert_logo_3,t.cert_background,t.cert_signature,t.cert_stamp
        FROM registrations r JOIN events e ON r.event_id=e.id JOIN tournaments t ON e.tournament_id=t.id
        WHERE r.event_id=? AND r.status='approved' AND r.is_waitlist=0 AND r.is_complete=1 ORDER BY r.id''',(event_id,)).fetchall()
    certificates=[]
    for reg in regs:
        members=conn.execute('SELECT * FROM registration_members WHERE registration_id=? ORDER BY id',(reg['id'],)).fetchall()
        for member in members:
            code=get_or_create_cert(conn,reg['id'],member['id'],'individual')
            certificates.append({'reg':reg,'member':member,'verification_code':code,'team_mode':False,'coach_mode':False})
        # เพิ่มเกียรติบัตรผู้ฝึกสอน 1 ใบต่อ 1 ทีม/รายการสมัคร
        if (reg['coach_name'] or '').strip():
            code=get_or_create_cert(conn,reg['id'],None,'coach')
            certificates.append({'reg':reg,'member':{'member_name':reg['coach_name']},'verification_code':code,'team_mode':False,'coach_mode':True})
    conn.close()
    if not certificates:
        flash('ยังไม่มีรายชื่อที่อนุมัติแล้วในอีเวนต์นี้'); return redirect(url_for('tournament_registrations',tournament_id=event['tournament_id'],event_id=event_id))
    return render_template('certificate_bulk.html',certificates=certificates,event=event)

@app.route('/certificate/<code>/qr.png')
def certificate_qr(code):
    conn=get_db(); cert=conn.execute('SELECT id FROM certificates WHERE verification_code=?',(code,)).fetchone(); conn.close()
    if not cert: abort(404)
    img=qrcode.make(url_for('verify_certificate',code=code,_external=True)); out=BytesIO(); img.save(out,format='PNG'); out.seek(0)
    return send_file(out,mimetype='image/png')

@app.route('/verify/<code>')
def verify_certificate(code):
    conn=get_db(); cert=conn.execute('''SELECT c.*,r.team_name,r.affiliation,r.coach_name,r.award_result,r.award_custom,m.member_name,e.event_name,e.category_type,e.gender_type,e.age_group,e.event_mode,e.sport_name,e.fixed_member_count,t.title tournament_title FROM certificates c JOIN registrations r ON c.registration_id=r.id LEFT JOIN registration_members m ON c.member_id=m.id JOIN events e ON r.event_id=e.id JOIN tournaments t ON e.tournament_id=t.id WHERE c.verification_code=?''',(code,)).fetchone(); conn.close(); return render_template('verify_certificate.html',cert=cert,code=code)

def safe_sheet_title(raw, used_titles):
    invalid='[]:*?/\\'
    title=''.join('_' if ch in invalid else ch for ch in str(raw or 'อีเวนต์')).strip() or 'อีเวนต์'
    title=title[:31]
    base=title
    suffix=2
    while title in used_titles:
        marker=f'_{suffix}'
        title=(base[:31-len(marker)] + marker)
        suffix += 1
    used_titles.add(title)
    return title


def append_registration_export_sheet(conn, ws, event, registrations):
    ws.append(['ลำดับ','รหัสสมัคร','อีเวนต์','ชื่อทีม','ต้นสังกัด','ผู้ติดต่อ','เบอร์โทร','จำนวนผู้เล่น','ข้อมูลครบ','สถานะ','ผลการแข่งขัน','สำรอง','แหล่งข้อมูล','สมาชิก 1','เลขบัตร 1','สมาชิก 2','เลขบัตร 2','สมาชิก 3','เลขบัตร 3','สมาชิก 4','เลขบัตร 4','หมายเหตุ'])
    for idx,r in enumerate(registrations,1):
        ms=conn.execute('SELECT * FROM registration_members WHERE registration_id=? ORDER BY id',(r['id'],)).fetchall()
        row=[idx,r['registration_code'],event_display_name(event),r['team_name'] or '',r['affiliation'] or '',r['contact_name'],r['phone'],r['member_count'],'ครบ' if r['is_complete'] else 'ยังไม่ครบ',r['status'],award_label(r['award_result'],r['award_custom']),'ใช่' if r['is_waitlist'] else 'ไม่',r['source']]
        for i in range(4): row += [ms[i]['member_name'],ms[i]['member_idcard'] or ''] if i<len(ms) else ['','']
        row += [r['notes'] or '']
        ws.append(row)
    style_ws(ws)


@app.route('/admin/tournament/<int:tournament_id>/export-all')
def export_tournament_excel(tournament_id):
    """Download one workbook for competition management, separated by event sheets."""
    if not is_logged_in(): return redirect(url_for('login'))
    conn=get_db(); tournament=get_owned_tournament(conn,tournament_id)
    if not tournament: conn.close(); abort(404)
    events=conn.execute('SELECT * FROM events WHERE tournament_id=? ORDER BY age_group,category_type,gender_type,id',(tournament_id,)).fetchall()
    wb=openpyxl.Workbook(); summary_ws=wb.active; summary_ws.title='สรุปอีเวนต์'
    summary_ws.append(['ลำดับ','อีเวนต์','ประเภท','เพศ','รุ่น','ผู้สมัครหลัก','Waiting List','ข้อมูลยังไม่ครบ','อนุมัติแล้ว'])
    used_titles={'สรุปอีเวนต์'}
    for idx,event in enumerate(events,1):
        regs=conn.execute('SELECT * FROM registrations WHERE event_id=? ORDER BY is_waitlist,id',(event['id'],)).fetchall()
        active=sum(1 for r in regs if not r['is_waitlist'])
        waiting=sum(1 for r in regs if r['is_waitlist'])
        incomplete=sum(1 for r in regs if not r['is_complete'])
        approved=sum(1 for r in regs if r['status']=='approved')
        summary_ws.append([idx,event_display_name(event),category_label(event['category_type']),gender_label(event['gender_type']),age_label(event['age_group']),active,waiting,incomplete,approved])
        ws=wb.create_sheet(safe_sheet_title(event_display_name(event),used_titles))
        append_registration_export_sheet(conn,ws,event,regs)
    conn.close(); style_ws(summary_ws)
    filename=f'tournament_{tournament_id}_all_registrations.xlsx'
    return excel_response(wb,filename)


@app.route('/admin/event/<int:event_id>/export')
def export_event_excel(event_id):
    if not is_logged_in(): return redirect(url_for('login'))
    conn=get_db(); e=get_owned_event(conn,event_id)
    if not e or e['created_by']!=session['user_id']: conn.close(); abort(404)
    regs=conn.execute('SELECT * FROM registrations WHERE event_id=? ORDER BY id',(event_id,)).fetchall(); wb=openpyxl.Workbook(); ws=wb.active; ws.title='รายชื่อสมัคร'; ws.append(['ลำดับ','รหัสสมัคร','อีเวนต์','ชื่อทีม','ต้นสังกัด','ผู้ติดต่อ','เบอร์โทร','จำนวนผู้เล่น','สถานะ','ผลการแข่งขัน','สำรอง','แหล่งข้อมูล','สมาชิก 1','เลขบัตร 1','สมาชิก 2','เลขบัตร 2','สมาชิก 3','เลขบัตร 3','สมาชิก 4','เลขบัตร 4','หมายเหตุ'])
    for idx,r in enumerate(regs,1):
        ms=conn.execute('SELECT * FROM registration_members WHERE registration_id=? ORDER BY id',(r['id'],)).fetchall(); row=[idx,r['registration_code'],event_display_name(e),r['team_name'] or '',r['affiliation'] or '',r['contact_name'],r['phone'],r['member_count'],r['status'],award_label(r['award_result'],r['award_custom']),'ใช่' if r['is_waitlist'] else 'ไม่',r['source']]
        for i in range(4): row += [ms[i]['member_name'],ms[i]['member_idcard'] or ''] if i<len(ms) else ['','']
        row += [r['notes'] or '']; ws.append(row)
    conn.close(); style_ws(ws); return excel_response(wb,f'event_{event_id}_registrations.xlsx')

def style_ws(ws):
    for cell in ws[1]: cell.font=Font(bold=True); cell.fill=PatternFill('solid',fgColor='DDEBFF'); cell.alignment=Alignment(horizontal='center')
    for col in ws.columns:
        letter=col[0].column_letter; ws.column_dimensions[letter].width=min(max(12,max(len(str(c.value or '')) for c in col)+2),35)
    ws.freeze_panes='A2'; ws.auto_filter.ref=ws.dimensions

def excel_response(wb,name):
    out=BytesIO(); wb.save(out); out.seek(0); return send_file(out,as_attachment=True,download_name=name,mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/admin/tournament/<int:tournament_id>/registrations/delete-bulk',methods=['POST'])
def delete_registrations_bulk(tournament_id):
    if not is_logged_in(): return redirect(url_for('login'))
    conn=get_db(); tournament=get_owned_tournament(conn,tournament_id)
    if not tournament:
        conn.close(); return redirect(url_for('admin_dashboard'))
    selected_event_id=request.form.get('event_id',type=int)
    status_filter=request.form.get('status_filter','pending')
    if status_filter not in {'pending','approved'}: status_filter='pending'
    if selected_event_id:
        owned_event=conn.execute('SELECT id FROM events WHERE id=? AND tournament_id=?',(selected_event_id,tournament_id)).fetchone()
        if not owned_event: selected_event_id=None
    action=request.form.get('action','delete_selected')
    ids=[]
    all_actions={'delete_all','approve_all'}
    if action in all_actions:
        query='SELECT r.id FROM registrations r JOIN events e ON r.event_id=e.id WHERE e.tournament_id=?'
        args=[tournament_id]
        if selected_event_id:
            query += ' AND e.id=?'; args.append(selected_event_id)
        query += " AND r.status='approved'" if status_filter=='approved' else " AND r.status!='approved'"
        ids=[r['id'] for r in conn.execute(query,args).fetchall()]
    else:
        raw_ids=request.form.getlist('registration_ids')
        selected=[]
        for raw in raw_ids:
            try: selected.append(int(raw))
            except (TypeError,ValueError): pass
        if selected:
            marks=','.join(['?']*len(selected))
            query=f'SELECT r.id FROM registrations r JOIN events e ON r.event_id=e.id WHERE e.tournament_id=? AND r.id IN ({marks})'
            ids=[r['id'] for r in conn.execute(query,(tournament_id,*selected)).fetchall()]

    redirect_url=url_for('tournament_registrations',tournament_id=tournament_id,event_id=selected_event_id,status_filter=status_filter) if selected_event_id else url_for('tournament_registrations',tournament_id=tournament_id,status_filter=status_filter)
    if not ids:
        conn.close()
        flash('ยังไม่ได้เลือกรายการ')
        return redirect(redirect_url)

    if action in {'approve_selected','approve_all'}:
        marks=','.join(['?']*len(ids))
        eligible=conn.execute(f'''SELECT id FROM registrations
            WHERE id IN ({marks}) AND is_complete=1 AND is_waitlist=0 AND status!='approved' ''',tuple(ids)).fetchall()
        eligible_ids=[r['id'] for r in eligible]
        if eligible_ids:
            eligible_marks=','.join(['?']*len(eligible_ids))
            conn.execute(f'''UPDATE registrations SET status='approved',approved_at=?
                WHERE id IN ({eligible_marks})''',(now(),*eligible_ids))
        skipped=len(ids)-len(eligible_ids)
        conn.commit(); conn.close()
        message=f'อนุมัติผู้สมัครเรียบร้อยแล้ว {len(eligible_ids)} รายการ'
        if skipped: message += f' · ข้าม {skipped} รายการที่อนุมัติไม่ได้หรืออนุมัติไว้แล้ว'
        flash(message)
        return redirect(redirect_url)

    conn.close()
    for rid in ids: delete_registration_internal(rid)
    flash(f'ลบผู้สมัครเรียบร้อยแล้ว {len(ids)} รายการ')
    return redirect(redirect_url)

@app.route('/admin/registration/<int:registration_id>/delete')
def delete_registration(registration_id):
    if not is_logged_in(): return redirect(url_for('login'))
    conn=get_db(); r=conn.execute('''SELECT r.*,e.tournament_id,t.created_by FROM registrations r JOIN events e ON r.event_id=e.id JOIN tournaments t ON e.tournament_id=t.id WHERE r.id=?''',(registration_id,)).fetchone()
    if not r or r['created_by']!=session['user_id']: conn.close(); return redirect(url_for('admin_dashboard'))
    for m in conn.execute('SELECT idcard_file FROM registration_members WHERE registration_id=?',(registration_id,)).fetchall(): delete_uploaded_file(m['idcard_file'])
    delete_uploaded_file(r['slip_filename']); conn.execute('DELETE FROM certificates WHERE registration_id=?',(registration_id,)); conn.execute('DELETE FROM registration_members WHERE registration_id=?',(registration_id,)); conn.execute('DELETE FROM registrations WHERE id=?',(registration_id,)); conn.commit(); conn.close(); flash('ลบผู้สมัครเรียบร้อยแล้ว'); return redirect(url_for('tournament_registrations',tournament_id=r['tournament_id'],event_id=r['event_id']))

@app.route('/admin/event/<int:event_id>/delete')
def delete_event(event_id):
    if not is_logged_in(): return redirect(url_for('login'))
    conn=get_db(); e=get_owned_event(conn,event_id)
    if not e or e['created_by']!=session['user_id']: conn.close(); return redirect(url_for('admin_dashboard'))
    regs=conn.execute('SELECT id FROM registrations WHERE event_id=?',(event_id,)).fetchall(); conn.close()
    for r in regs: delete_registration_internal(r['id'])
    conn=get_db(); conn.execute('DELETE FROM events WHERE id=?',(event_id,)); conn.commit(); conn.close(); flash('ลบอีเวนต์เรียบร้อยแล้ว'); return redirect(url_for('manage_events',tournament_id=e['tournament_id']))

def delete_registration_internal(rid):
    conn=get_db(); r=conn.execute('SELECT * FROM registrations WHERE id=?',(rid,)).fetchone()
    if r:
        for m in conn.execute('SELECT idcard_file FROM registration_members WHERE registration_id=?',(rid,)).fetchall(): delete_uploaded_file(m['idcard_file'])
        delete_uploaded_file(r['slip_filename']); conn.execute('DELETE FROM certificates WHERE registration_id=?',(rid,)); conn.execute('DELETE FROM registration_members WHERE registration_id=?',(rid,)); conn.execute('DELETE FROM registrations WHERE id=?',(rid,)); conn.commit()
    conn.close()

@app.route('/admin/tournament/<int:tournament_id>/delete')
def delete_tournament(tournament_id):
    if not is_logged_in(): return redirect(url_for('login'))
    conn=get_db(); t=get_owned_tournament(conn,tournament_id)
    if not t: conn.close(); return redirect(url_for('admin_dashboard'))
    eventids=[e['id'] for e in conn.execute('SELECT id FROM events WHERE tournament_id=?',(tournament_id,)).fetchall()]; conn.close()
    for eid in eventids:
        conn=get_db(); regs=conn.execute('SELECT id FROM registrations WHERE event_id=?',(eid,)).fetchall(); conn.close()
        for r in regs: delete_registration_internal(r['id'])
    conn=get_db(); conn.execute('DELETE FROM events WHERE tournament_id=?',(tournament_id,)); conn.execute('DELETE FROM tournaments WHERE id=?',(tournament_id,)); conn.commit(); conn.close(); flash('ลบงานแข่งขันเรียบร้อยแล้ว'); return redirect(url_for('admin_dashboard'))

@app.route('/health')
def health():
    conn=get_db(); conn.execute('SELECT 1').fetchone(); conn.close()
    return jsonify(status='ok', database='postgresql' if IS_POSTGRES else 'sqlite')

init_db()

if __name__=='__main__':
    app.run(host='0.0.0.0',port=int(os.environ.get('PORT',8000)),debug=True)
